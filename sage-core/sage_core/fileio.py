"""SAGE Core — Excel/CSV File I/O.

Handles reading enterprise data files (Excel, CSV) into DataFrames and writing
optimization results back to formatted Excel workbooks.

Public API
----------
read_data(filepath) -> dict[str, pd.DataFrame]
read_data_from_bytes(content, filename) -> dict[str, pd.DataFrame]
write_results_excel(result, model_name, output_path, original_data) -> str
write_results_csv(result, output_path) -> str
generate_template(problem_type, output_path) -> str
dataframe_to_model(dfs, problem_type) -> LPModel | MIPModel | PortfolioModel | SchedulingModel

Design constraints
------------------
- No filesystem access beyond what callers explicitly pass as paths/bytes.
- No print() calls.
- No global state.
- All functions are pure (take in → return out) except for file writes,
  which create/overwrite the path provided by the caller.
"""

from __future__ import annotations

import io
import logging
import os
import re
from typing import TYPE_CHECKING

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from sage_core.models import (
    Asset,
    DataValidationError,
    FileIOError,
    LPModel,
    LPVariable,
    LinearConstraint,
    LinearObjective,
    MIPModel,
    MIPVariable,
    PortfolioConstraints,
    PortfolioModel,
    SchedulingModel,
    Shift,
    SolverResult,
    Worker,
)

if TYPE_CHECKING:
    pass

logger = logging.getLogger("sage.fileio")

# ---------------------------------------------------------------------------
# Colour / style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")   # blue
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FONT = Font(italic=True, color="808080", size=10)
_BINDING_FILL = PatternFill("solid", fgColor="FFF2CC")   # pale yellow
_EXAMPLE_FILL = PatternFill("solid", fgColor="E8F0FE")   # pale blue (example rows)
_THIN_BORDER_SIDE = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)

_SUPPORTED_PROBLEM_TYPES = {"portfolio", "scheduling", "transport", "generic_lp"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _clean_string(value: object) -> str:
    """Coerce to str and strip surrounding whitespace."""
    return str(value).strip() if value is not None else ""


def _parse_number(raw: object, context: str = "") -> float:
    """Convert messy number representations to float.

    Handles:
    - Plain floats/ints
    - Percentage strings ("5%" → 0.05, "5.2 %" → 0.052)
    - Comma-separated thousands ("1,000.50" → 1000.5)
    - Leading/trailing whitespace

    Raises:
        ValueError: If the value cannot be parsed as a number.
    """
    if isinstance(raw, (int, float)):
        return float(raw)
    s = _clean_string(raw)
    if not s or s.lower() in ("nan", "none", "n/a", ""):
        raise ValueError(f"Empty or null value{' in ' + context if context else ''}")
    # Percentage
    is_pct = s.endswith("%")
    s_num = s.rstrip("% ").replace(",", "")
    try:
        val = float(s_num)
    except ValueError:
        raise ValueError(
            f"Cannot parse {raw!r} as a number{' in ' + context if context else ''}"
        )
    return val / 100.0 if is_pct else val


def _parse_optional_number(raw: object, default: float | None = None) -> float | None:
    """Parse a number that may be blank/NaN → default."""
    if raw is None or (isinstance(raw, float) and np.isnan(raw)):
        return default
    s = _clean_string(raw)
    if not s or s.lower() in ("nan", "none", "n/a", ""):
        return default
    try:
        return _parse_number(raw)
    except ValueError:
        return default


def _auto_width(ws: "openpyxl.worksheet.worksheet.Worksheet") -> None:
    """Set column widths based on maximum content length + padding."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


def _style_header_row(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    row_idx: int,
    n_cols: int,
) -> None:
    """Apply header fill/font/border to a row in a worksheet."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_row(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    row_idx: int,
    values: list,
    *,
    bold: bool = False,
    fill: PatternFill | None = None,
    number_format: str = "General",
) -> None:
    """Write a list of values to a worksheet row."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
        if number_format != "General":
            cell.number_format = number_format
        cell.border = _THIN_BORDER


def _strip_blank(df: pd.DataFrame) -> pd.DataFrame:
    """Remove fully blank rows and columns from a DataFrame."""
    df = df.dropna(how="all").dropna(axis=1, how="all")
    return df.reset_index(drop=True)


def _forward_fill_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Forward-fill NaN values (handles merged cells from Excel read)."""
    return df.ffill()


# ---------------------------------------------------------------------------
# Public API — reading
# ---------------------------------------------------------------------------


def read_data(filepath: str, file_format: str = "auto") -> dict[str, pd.DataFrame]:
    """Read an Excel or CSV file into DataFrames, one per sheet.

    Auto-detects file format from the extension unless ``file_format`` is
    specified.  Applies basic cleaning: blank row/column stripping and
    merged-cell forward fill.

    Args:
        filepath: Absolute or relative path to the file.
        file_format: ``"auto"``, ``"excel"``, ``"csv"``, or ``"tsv"``.

    Returns:
        Dict mapping sheet name (or ``"data"`` for CSV) to DataFrame.

    Raises:
        FileIOError: If the file cannot be read or the format is unsupported.
    """
    if not os.path.isfile(filepath):
        raise FileIOError(
            f"File not found: {filepath}",
            details={"filepath": filepath},
            suggestions=[
                "Check the file path is correct",
                "Ensure the file exists and is readable",
            ],
        )
    try:
        with open(filepath, "rb") as fh:
            content = fh.read()
    except OSError as exc:
        raise FileIOError(
            f"Cannot read file '{filepath}': {exc}",
            details={"filepath": filepath, "os_error": str(exc)},
        ) from exc

    return read_data_from_bytes(content, os.path.basename(filepath), file_format)


def read_data_from_bytes(
    content: bytes,
    filename: str,
    file_format: str = "auto",
) -> dict[str, pd.DataFrame]:
    """Read Excel or CSV bytes into DataFrames.

    Identical logic to :func:`read_data` but accepts an in-memory bytes buffer.
    Useful for cloud deployments where files arrive via HTTP upload.

    Args:
        content: Raw file bytes.
        filename: Original filename (used for format detection).
        file_format: ``"auto"``, ``"excel"``, ``"csv"``, or ``"tsv"``.

    Returns:
        Dict mapping sheet name (or ``"data"`` for CSV/TSV) to DataFrame.

    Raises:
        FileIOError: If the format is unsupported or parsing fails.
    """
    ext = os.path.splitext(filename)[1].lower()

    if file_format == "auto":
        if ext in (".xlsx", ".xls", ".xlsm"):
            file_format = "excel"
        elif ext == ".csv":
            file_format = "csv"
        elif ext in (".tsv", ".tab"):
            file_format = "tsv"
        else:
            raise FileIOError(
                f"Cannot auto-detect format for '{filename}'. "
                f"Supported extensions: .xlsx, .xls, .csv, .tsv",
                details={"filename": filename, "extension": ext},
                suggestions=["Rename the file with a supported extension"],
            )

    if file_format == "excel":
        return _read_excel_bytes(content, filename)
    elif file_format in ("csv", "tsv"):
        sep = "\t" if file_format == "tsv" else ","
        return {"data": _read_csv_bytes(content, filename, sep)}
    else:
        raise FileIOError(
            f"Unsupported file_format: '{file_format}'",
            details={"file_format": file_format},
            suggestions=["Use 'auto', 'excel', 'csv', or 'tsv'"],
        )


def _read_excel_bytes(content: bytes, filename: str) -> dict[str, pd.DataFrame]:
    """Read all sheets from Excel bytes."""
    buf = io.BytesIO(content)
    try:
        # Use openpyxl with data_only=True to resolve formula cells
        wb = openpyxl.load_workbook(buf, data_only=True)
    except Exception as exc:
        raise FileIOError(
            f"Cannot parse Excel file '{filename}': {exc}",
            details={"filename": filename, "error": str(exc)},
            suggestions=[
                "Ensure the file is a valid .xlsx/.xls workbook",
                "Try re-saving the file in Excel and try again",
            ],
        ) from exc

    result: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = pd.DataFrame()
            continue

        # Find the first non-blank row to use as header
        header_idx = 0
        for i, row in enumerate(rows):
            non_null = [c for c in row if c is not None and str(c).strip() != ""]
            if non_null:
                header_idx = i
                break

        header = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(rows[header_idx])]
        data_rows = rows[header_idx + 1 :]

        df = pd.DataFrame(data_rows, columns=header)
        df = _forward_fill_headers(df)
        df = _strip_blank(df)
        result[sheet_name] = df

    return result


def _read_csv_bytes(content: bytes, filename: str, sep: str = ",") -> pd.DataFrame:
    """Read CSV bytes with encoding auto-detection."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(encoding)
            df = pd.read_csv(io.StringIO(text), sep=sep)
            df = _forward_fill_headers(df)
            df = _strip_blank(df)
            logger.debug("CSV '%s' decoded with %s, %d rows", filename, encoding, len(df))
            return df
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            raise FileIOError(
                f"Cannot parse CSV '{filename}': {exc}",
                details={"filename": filename, "error": str(exc)},
            ) from exc
    raise FileIOError(
        f"Cannot decode CSV '{filename}' — tried utf-8, latin-1, cp1252",
        details={"filename": filename},
        suggestions=["Save the file as UTF-8 encoding and try again"],
    )


# ---------------------------------------------------------------------------
# Public API — writing results
# ---------------------------------------------------------------------------


def write_results_excel(
    result: SolverResult,
    model_name: str,
    output_path: str,
    original_data: dict | None = None,
) -> str:
    """Write optimization results to a formatted multi-sheet Excel workbook.

    Sheets created:
    - **Summary**: status, objective, solve time, variable/constraint counts.
    - **Solution**: variable names and optimal values, sorted by value descending.
    - **Sensitivity** (LP optimal only): shadow prices, reduced costs, ranges.
    - **Constraints** (optimal only): name, RHS, slack, binding status.
    - **Infeasibility** (infeasible only): IIS constraints and explanation.

    Formatting applied:
    - Header row: bold white text on blue fill (#2E75B6), frozen.
    - Numbers: 4 decimal places for values, 2 for percentages.
    - Column auto-width.
    - Binding constraints highlighted in yellow.

    Args:
        result: Solver result to write.
        model_name: Human-readable model name for the Summary sheet.
        output_path: Destination ``.xlsx`` path. Created or overwritten.
        original_data: Unused (reserved for future passthrough).

    Returns:
        The ``output_path`` string.

    Raises:
        FileIOError: If the workbook cannot be saved.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_summary_sheet(wb, result, model_name)

    if result.status in ("optimal", "time_limit_reached") and result.variable_values:
        _write_solution_sheet(wb, result)
        _write_constraints_sheet(wb, result)
        if result.shadow_prices:
            _write_sensitivity_sheet(wb, result)

    if result.status == "infeasible" and result.iis:
        _write_infeasibility_sheet(wb, result)

    try:
        wb.save(output_path)
    except OSError as exc:
        raise FileIOError(
            f"Cannot save results to '{output_path}': {exc}",
            details={"output_path": output_path, "os_error": str(exc)},
            suggestions=["Check the directory exists and you have write permission"],
        ) from exc

    logger.debug("Results written to '%s'", output_path)
    return output_path


def _write_summary_sheet(
    wb: openpyxl.Workbook,
    result: SolverResult,
    model_name: str,
) -> None:
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "B2"

    headers = ["Property", "Value"]
    _style_header_row(ws, 1, len(headers))
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    rows_data = [
        ("Model name", model_name),
        ("Status", result.status),
        (
            "Objective value",
            f"{result.objective_value:.6g}" if result.objective_value is not None else "—",
        ),
        ("Solve time (s)", f"{result.solve_time_seconds:.4f}"),
    ]
    if result.bound is not None:
        rows_data.append(("Best bound", f"{result.bound:.6g}"))
    if result.gap is not None:
        rows_data.append(("MIP gap", f"{result.gap:.4%}"))
    if result.variable_values:
        rows_data.append(("Variables", str(len(result.variable_values))))

    for r, (prop, val) in enumerate(rows_data, start=2):
        ws.cell(row=r, column=1, value=prop).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        for col in (1, 2):
            ws.cell(row=r, column=col).border = _THIN_BORDER

    if result.iis:
        r += 1
        ws.cell(row=r + 1, column=1, value="IIS explanation").font = Font(bold=True)
        ws.cell(row=r + 1, column=2, value=result.iis.explanation)
        ws.cell(row=r + 1, column=2).alignment = Alignment(wrap_text=True)

    _auto_width(ws)


def _write_solution_sheet(wb: openpyxl.Workbook, result: SolverResult) -> None:
    ws = wb.create_sheet("Solution")
    ws.freeze_panes = "A2"

    headers = ["Variable", "Optimal Value"]
    _style_header_row(ws, 1, len(headers))
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    sorted_vars = sorted(
        result.variable_values.items(), key=lambda kv: kv[1], reverse=True
    )
    for r, (name, val) in enumerate(sorted_vars, start=2):
        ws.cell(row=r, column=1, value=name).border = _THIN_BORDER
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = "0.0000"
        cell.border = _THIN_BORDER

    _auto_width(ws)


def _write_sensitivity_sheet(wb: openpyxl.Workbook, result: SolverResult) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws.freeze_panes = "A2"
    row = 1

    # Shadow prices
    if result.shadow_prices:
        hdrs = ["Constraint", "Shadow Price"]
        if result.rhs_ranges:
            hdrs += ["RHS Range Lower", "RHS Range Upper"]
        _style_header_row(ws, row, len(hdrs))
        for col, h in enumerate(hdrs, start=1):
            ws.cell(row=row, column=col, value=h)
        row += 1
        for cname, sp in result.shadow_prices.items():
            ws.cell(row=row, column=1, value=cname).border = _THIN_BORDER
            c2 = ws.cell(row=row, column=2, value=sp)
            c2.number_format = "0.0000"
            c2.border = _THIN_BORDER
            if result.rhs_ranges and cname in result.rhs_ranges:
                lo, hi = result.rhs_ranges[cname]
                c3 = ws.cell(row=row, column=3, value=lo)
                c3.number_format = "0.0000"
                c3.border = _THIN_BORDER
                c4 = ws.cell(row=row, column=4, value=hi)
                c4.number_format = "0.0000"
                c4.border = _THIN_BORDER
            row += 1
        row += 1  # blank separator

    # Reduced costs
    if result.reduced_costs:
        hdrs = ["Variable", "Reduced Cost"]
        if result.objective_ranges:
            hdrs += ["Obj Range Lower", "Obj Range Upper"]
        _style_header_row(ws, row, len(hdrs))
        for col, h in enumerate(hdrs, start=1):
            ws.cell(row=row, column=col, value=h)
        row += 1
        for vname, rc in result.reduced_costs.items():
            ws.cell(row=row, column=1, value=vname).border = _THIN_BORDER
            c2 = ws.cell(row=row, column=2, value=rc)
            c2.number_format = "0.0000"
            c2.border = _THIN_BORDER
            if result.objective_ranges and vname in result.objective_ranges:
                lo, hi = result.objective_ranges[vname]
                c3 = ws.cell(row=row, column=3, value=lo)
                c3.number_format = "0.0000"
                c3.border = _THIN_BORDER
                c4 = ws.cell(row=row, column=4, value=hi)
                c4.number_format = "0.0000"
                c4.border = _THIN_BORDER
            row += 1

    _auto_width(ws)


def _write_constraints_sheet(wb: openpyxl.Workbook, result: SolverResult) -> None:
    if not result.constraint_slack:
        return
    ws = wb.create_sheet("Constraints")
    ws.freeze_panes = "A2"

    headers = ["Constraint", "Slack", "Binding"]
    _style_header_row(ws, 1, len(headers))
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    binding_set = set(result.binding_constraints or [])
    for r, (cname, slack) in enumerate(result.constraint_slack.items(), start=2):
        is_binding = cname in binding_set
        fill = _BINDING_FILL if is_binding else None

        c1 = ws.cell(row=r, column=1, value=cname)
        c1.border = _THIN_BORDER
        if fill:
            c1.fill = fill

        c2 = ws.cell(row=r, column=2, value=slack)
        c2.number_format = "0.0000"
        c2.border = _THIN_BORDER
        if fill:
            c2.fill = fill

        c3 = ws.cell(row=r, column=3, value="Yes" if is_binding else "No")
        c3.border = _THIN_BORDER
        if fill:
            c3.fill = fill

    _auto_width(ws)


def _write_infeasibility_sheet(wb: openpyxl.Workbook, result: SolverResult) -> None:
    ws = wb.create_sheet("Infeasibility")
    iis = result.iis

    ws.cell(row=1, column=1, value="IIS — Irreducible Infeasible Subsystem").font = Font(
        bold=True, size=12
    )
    ws.cell(row=2, column=1, value=iis.explanation)
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 60

    row = 4
    headers = ["Conflicting Constraint"]
    _style_header_row(ws, row, len(headers))
    ws.cell(row=row, column=1, value=headers[0])
    row += 1
    for cname in iis.conflicting_constraints:
        c = ws.cell(row=row, column=1, value=cname)
        c.border = _THIN_BORDER
        c.fill = _BINDING_FILL
        row += 1

    if iis.conflicting_variable_bounds:
        row += 1
        ws.cell(row=row, column=1, value="Conflicting Variable Bounds").font = Font(bold=True)
        row += 1
        for bd in iis.conflicting_variable_bounds:
            c = ws.cell(row=row, column=1, value=bd)
            c.border = _THIN_BORDER
            row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 50


def write_results_csv(result: SolverResult, output_path: str) -> str:
    """Write variable values to a simple CSV.

    One row per variable with columns ``variable`` and ``value``.
    If no solution exists, writes an empty CSV with a status column.

    Args:
        result: Solver result.
        output_path: Destination ``.csv`` path.

    Returns:
        The ``output_path`` string.

    Raises:
        FileIOError: If the file cannot be written.
    """
    try:
        if result.variable_values:
            df = pd.DataFrame(
                [
                    {"variable": k, "value": v}
                    for k, v in sorted(result.variable_values.items())
                ]
            )
        else:
            df = pd.DataFrame([{"status": result.status, "variable": None, "value": None}])
        df.to_csv(output_path, index=False)
    except OSError as exc:
        raise FileIOError(
            f"Cannot write CSV to '{output_path}': {exc}",
            details={"output_path": output_path, "os_error": str(exc)},
        ) from exc
    return output_path


# ---------------------------------------------------------------------------
# Public API — template generation
# ---------------------------------------------------------------------------


def generate_template(problem_type: str, output_path: str) -> str:
    """Generate a blank Excel template for a given problem type.

    Creates a professional workbook with:
    - An **Instructions** sheet explaining what to fill in.
    - Data sheets with column headers, description row (italic gray), and
      3-5 example rows (pale-blue fill, clearly marked).
    - Data validation where appropriate.

    Supported problem types: ``portfolio``, ``scheduling``, ``transport``,
    ``generic_lp``.

    Args:
        problem_type: One of the supported types.
        output_path: Destination ``.xlsx`` path.

    Returns:
        The ``output_path`` string.

    Raises:
        FileIOError: If the workbook cannot be saved.
        DataValidationError: If ``problem_type`` is not recognised.
    """
    if problem_type not in _SUPPORTED_PROBLEM_TYPES:
        raise DataValidationError(
            f"Unknown problem type: '{problem_type}'. "
            f"Supported: {sorted(_SUPPORTED_PROBLEM_TYPES)}",
            details={"problem_type": problem_type},
            suggestions=[f"Use one of: {', '.join(sorted(_SUPPORTED_PROBLEM_TYPES))}"],
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    generators = {
        "portfolio": _gen_portfolio_template,
        "scheduling": _gen_scheduling_template,
        "transport": _gen_transport_template,
        "generic_lp": _gen_generic_lp_template,
    }
    generators[problem_type](wb)

    try:
        wb.save(output_path)
    except OSError as exc:
        raise FileIOError(
            f"Cannot save template to '{output_path}': {exc}",
            details={"output_path": output_path, "os_error": str(exc)},
        ) from exc

    logger.debug("Template '%s' written to '%s'", problem_type, output_path)
    return output_path


def _add_instructions(wb: openpyxl.Workbook, text_lines: list[str]) -> None:
    """Add an Instructions sheet as the first sheet."""
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 80
    ws.cell(row=1, column=1, value="INSTRUCTIONS").font = Font(
        bold=True, size=14, color="2E75B6"
    )
    for i, line in enumerate(text_lines, start=2):
        c = ws.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 18


def _add_data_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    headers: list[str],
    descriptions: list[str],
    examples: list[list],
) -> "openpyxl.worksheet.worksheet.Worksheet":
    """Add a data sheet with headers, description row, and example rows."""
    ws = wb.create_sheet(sheet_name)
    n = len(headers)

    # Row 1: headers
    _style_header_row(ws, 1, n)
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.freeze_panes = "A2"

    # Row 2: descriptions (italic, gray)
    for col, desc in enumerate(descriptions, start=1):
        c = ws.cell(row=2, column=col, value=desc)
        c.font = _SUBHEADER_FONT
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(wrap_text=True)
        c.border = _THIN_BORDER
        ws.row_dimensions[2].height = 30

    # Rows 3+: example data
    for row_offset, example_row in enumerate(examples):
        r = row_offset + 3
        for col, val in enumerate(example_row, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = _EXAMPLE_FILL
            c.border = _THIN_BORDER

    _auto_width(ws)
    return ws


def _gen_portfolio_template(wb: openpyxl.Workbook) -> None:
    _add_instructions(wb, [
        "PORTFOLIO OPTIMIZATION TEMPLATE",
        "",
        "How to use this template:",
        "1. Fill in the 'Assets' sheet with your asset data (name, expected return, sector).",
        "2. Fill in the 'Covariance' sheet with the n×n covariance matrix.",
        "   - Row 1 and Column A must contain asset names in the same order as the Assets sheet.",
        "3. Fill in the 'Constraints' sheet with portfolio constraints.",
        "4. Delete the example rows (rows 3-5 in each sheet) before uploading.",
        "",
        "Expected return should be a decimal (e.g. 0.08 for 8%) or a percentage (e.g. 8%).",
        "Covariance values should be in decimal form (not percentages).",
    ])

    _add_data_sheet(
        wb, "Assets",
        headers=["Name", "Expected Return", "Sector"],
        descriptions=["Asset identifier (unique)", "Annual return (decimal or %)", "Sector label (optional)"],
        examples=[
            ["US Equity", 0.09, "Equity"],
            ["EU Bonds", "4%", "Fixed Income"],
            ["Gold", 0.06, "Commodity"],
        ],
    )

    # Covariance sheet (3×3 example)
    ws_cov = wb.create_sheet("Covariance")
    ws_cov.freeze_panes = "B2"
    assets = ["US Equity", "EU Bonds", "Gold"]
    cov_vals = [
        [0.04, 0.002, 0.001],
        [0.002, 0.001, 0.0005],
        [0.001, 0.0005, 0.005],
    ]
    # Header row
    _style_header_row(ws_cov, 1, len(assets) + 1)
    ws_cov.cell(row=1, column=1, value="Asset \\ Asset")
    for j, name in enumerate(assets, start=2):
        ws_cov.cell(row=1, column=j, value=name)
    ws_cov.cell(row=1, column=1).fill = _HEADER_FILL
    ws_cov.cell(row=1, column=1).font = _HEADER_FONT
    for i, name in enumerate(assets):
        r = i + 2
        ws_cov.cell(row=r, column=1, value=name).font = Font(bold=True)
        ws_cov.cell(row=r, column=1).border = _THIN_BORDER
        for j, val in enumerate(cov_vals[i], start=2):
            c = ws_cov.cell(row=r, column=j, value=val)
            c.fill = _EXAMPLE_FILL
            c.border = _THIN_BORDER
    _auto_width(ws_cov)

    _add_data_sheet(
        wb, "Constraints",
        headers=["Parameter", "Value"],
        descriptions=["Constraint name", "Value (decimal or %)"],
        examples=[
            ["min_total_allocation", 1.0],
            ["max_total_allocation", 1.0],
            ["min_allocation_per_asset", 0.05],
            ["max_allocation_per_asset", 0.40],
            ["risk_aversion", 2.0],
        ],
    )


def _gen_scheduling_template(wb: openpyxl.Workbook) -> None:
    _add_instructions(wb, [
        "NURSE / WORKER SCHEDULING TEMPLATE",
        "",
        "How to use this template:",
        "1. Fill in the 'Workers' sheet with each worker's name, max hours, skills, and unavailability.",
        "2. Fill in the 'Shifts' sheet with each shift's name, duration, required workers, and required skills.",
        "3. Fill in the 'Constraints' sheet with scheduling parameters.",
        "4. Skills must exactly match between Workers and Shifts sheets.",
        "5. Delete the example rows before uploading.",
        "",
        "Unavailable_Shifts: comma-separated shift names the worker cannot do.",
        "Required_Skills: comma-separated skills all workers on this shift must have.",
    ])

    _add_data_sheet(
        wb, "Workers",
        headers=["Name", "Max_Hours", "Skills", "Unavailable_Shifts"],
        descriptions=["Worker name (unique)", "Max weekly hours", "Comma-separated skills", "Comma-separated unavailable shifts (optional)"],
        examples=[
            ["Alice", 40, "ICU,General", ""],
            ["Bob", 40, "ER,General", "Night"],
            ["Carol", 32, "ICU,ER", ""],
        ],
    )

    _add_data_sheet(
        wb, "Shifts",
        headers=["Name", "Duration_Hours", "Required_Workers", "Required_Skills"],
        descriptions=["Shift name (unique)", "Hours per shift", "Min workers needed", "Comma-separated required skills (optional)"],
        examples=[
            ["Morning", 8, 2, "General"],
            ["Afternoon", 8, 2, ""],
            ["Night", 8, 1, ""],
        ],
    )

    _add_data_sheet(
        wb, "Constraints",
        headers=["Parameter", "Value"],
        descriptions=["Parameter name", "Value"],
        examples=[
            ["planning_horizon_days", 7],
            ["max_consecutive_days", 5],
        ],
    )


def _gen_transport_template(wb: openpyxl.Workbook) -> None:
    _add_instructions(wb, [
        "TRANSPORTATION PROBLEM TEMPLATE",
        "",
        "How to use this template:",
        "1. 'Origins' sheet: supply points with their supply capacity.",
        "2. 'Destinations' sheet: demand points with their demand.",
        "3. 'Costs' sheet: unit shipping costs from each origin to each destination.",
        "   - Row 1 header and Column A must list the same origins/destinations.",
        "4. 'Supply_Demand' sheet: summary totals (auto-check).",
        "5. Delete the example rows before uploading.",
    ])

    _add_data_sheet(
        wb, "Origins",
        headers=["Name", "Supply"],
        descriptions=["Origin/warehouse name", "Available supply units"],
        examples=[
            ["Warehouse_A", 300],
            ["Warehouse_B", 400],
            ["Warehouse_C", 500],
        ],
    )

    _add_data_sheet(
        wb, "Destinations",
        headers=["Name", "Demand"],
        descriptions=["Destination/customer name", "Required demand units"],
        examples=[
            ["Customer_1", 250],
            ["Customer_2", 350],
            ["Customer_3", 400],
        ],
    )

    # Costs sheet (origin × destination matrix)
    ws_costs = wb.create_sheet("Costs")
    origins = ["Warehouse_A", "Warehouse_B", "Warehouse_C"]
    dests = ["Customer_1", "Customer_2", "Customer_3"]
    costs = [[2, 3, 1], [5, 4, 8], [5, 6, 8]]
    _style_header_row(ws_costs, 1, len(dests) + 1)
    ws_costs.cell(row=1, column=1, value="Origin \\ Destination")
    for j, d in enumerate(dests, start=2):
        ws_costs.cell(row=1, column=j, value=d)
    for i, o in enumerate(origins):
        r = i + 2
        ws_costs.cell(row=r, column=1, value=o).font = Font(bold=True)
        ws_costs.cell(row=r, column=1).border = _THIN_BORDER
        for j, c_val in enumerate(costs[i], start=2):
            c = ws_costs.cell(row=r, column=j, value=c_val)
            c.fill = _EXAMPLE_FILL
            c.border = _THIN_BORDER
    _auto_width(ws_costs)

    _add_data_sheet(
        wb, "Supply_Demand",
        headers=["Total Supply", "Total Demand", "Balanced"],
        descriptions=["Sum of all supply", "Sum of all demand", "Supply >= Demand?"],
        examples=[[1200, 1000, "Yes"]],
    )


def _gen_generic_lp_template(wb: openpyxl.Workbook) -> None:
    _add_instructions(wb, [
        "GENERIC LINEAR PROGRAM TEMPLATE",
        "",
        "How to use this template:",
        "1. 'Variables' sheet: define decision variables (name, lower bound, upper bound, type).",
        "2. 'Constraints' sheet: define each constraint row.",
        "   - Coefficients column: JSON-style dict, e.g. {\"x1\": 2, \"x2\": 1}",
        "3. 'Objective' sheet: define the objective function.",
        "4. Delete example rows before uploading.",
        "",
        "Variable type: 'continuous', 'integer', or 'binary'.",
        "Constraint sense: '<=', '>=', or '=='.",
    ])

    _add_data_sheet(
        wb, "Variables",
        headers=["Name", "Lower_Bound", "Upper_Bound", "Type"],
        descriptions=["Variable name (unique)", "Lower bound (default 0)", "Upper bound (blank = unlimited)", "continuous / integer / binary"],
        examples=[
            ["x1", 0, "", "continuous"],
            ["x2", 0, 100, "integer"],
            ["y1", 0, 1, "binary"],
        ],
    )

    _add_data_sheet(
        wb, "Constraints",
        headers=["Name", "Coefficients", "Sense", "RHS"],
        descriptions=["Constraint name", 'JSON dict e.g. {"x1": 2, "x2": 1}', "<= / >= / ==", "Right-hand side value"],
        examples=[
            ["resource_A", '{"x1": 2, "x2": 1}', "<=", 14],
            ["resource_B", '{"x1": 1, "x2": 2}', "<=", 14],
            ["min_x1", '{"x1": 1}', ">=", 1],
        ],
    )

    _add_data_sheet(
        wb, "Objective",
        headers=["Sense", "Coefficients"],
        descriptions=["minimize or maximize", 'JSON dict e.g. {"x1": 5, "x2": 4}'],
        examples=[["maximize", '{"x1": 5, "x2": 4}']],
    )


# ---------------------------------------------------------------------------
# Public API — DataFrame to model
# ---------------------------------------------------------------------------


def dataframe_to_model(
    dfs: dict[str, pd.DataFrame],
    problem_type: str,
) -> "LPModel | MIPModel | PortfolioModel | SchedulingModel":
    """Parse DataFrames (from Excel/CSV) into a typed Pydantic model.

    This is the critical bridge between file data and the solver pipeline.

    **Messy data handling:**
    - Extra columns are silently ignored.
    - Column names are stripped of whitespace and matched case-insensitively.
    - Missing optional columns use model defaults.
    - ``"1,000.50"`` → ``1000.5``; ``"5%"`` → ``0.05``.
    - Blank/NaN rows are skipped.

    Args:
        dfs: Dict of DataFrames, typically from :func:`read_data`.
        problem_type: One of ``"portfolio"``, ``"scheduling"``,
            ``"transport"``, ``"generic_lp"``.

    Returns:
        A validated Pydantic model instance.

    Raises:
        DataValidationError: If required data is missing or malformed.
            Includes sheet name, column, and row information.
    """
    if problem_type not in _SUPPORTED_PROBLEM_TYPES:
        raise DataValidationError(
            f"Unknown problem type: '{problem_type}'",
            details={"problem_type": problem_type},
            suggestions=[f"Use one of: {', '.join(sorted(_SUPPORTED_PROBLEM_TYPES))}"],
        )

    parsers = {
        "portfolio": _parse_portfolio,
        "scheduling": _parse_scheduling,
        "transport": _parse_transport_lp,
        "generic_lp": _parse_generic_lp,
    }
    return parsers[problem_type](dfs)


# ---------------------------------------------------------------------------
# Normalisation helpers for DataFrames
# ---------------------------------------------------------------------------


def _normalise_cols(df: pd.DataFrame) -> dict[str, str]:
    """Return mapping from normalised column name → actual column name."""
    return {str(c).strip().lower().replace(" ", "_"): str(c) for c in df.columns}


def _get_col(
    df: pd.DataFrame,
    col_norm: dict[str, str],
    key: str,
    *,
    sheet: str,
    required: bool = True,
) -> str | None:
    """Resolve a column by normalised key. Returns actual column name or None."""
    actual = col_norm.get(key)
    if actual is None and required:
        raise DataValidationError(
            f"Required column '{key}' not found in sheet '{sheet}'",
            details={
                "sheet": sheet,
                "required_column": key,
                "found_columns": list(df.columns),
                "missing": [key],
            },
            suggestions=[
                f"Add a column named '{key}' to the '{sheet}' sheet",
                f"Available columns: {list(df.columns)}",
            ],
        )
    return actual


def _parse_str_list(raw: object) -> list[str]:
    """Parse comma-separated string list. Returns [] for blank/null."""
    if raw is None or (isinstance(raw, float) and np.isnan(raw)):
        return []
    s = _clean_string(raw)
    if not s:
        return []
    return [x.strip() for x in s.split(",") if x.strip()]


# ---------------------------------------------------------------------------
# Portfolio parser
# ---------------------------------------------------------------------------


def _parse_portfolio(dfs: dict[str, pd.DataFrame]) -> PortfolioModel:
    # Locate sheets (case-insensitive)
    sheets_norm = {k.strip().lower(): k for k in dfs}

    def _get_sheet(key: str) -> pd.DataFrame:
        actual = sheets_norm.get(key)
        if actual is None:
            raise DataValidationError(
                f"Required sheet '{key}' not found",
                details={"found_sheets": list(dfs.keys()), "required": key},
                suggestions=[f"Add a sheet named '{key}' to the workbook"],
            )
        return dfs[actual]

    assets_df = _get_sheet("assets")
    cov_df = _get_sheet("covariance")
    constraints_df = _get_sheet("constraints") if "constraints" in sheets_norm else None

    # ---- Parse assets --------------------------------------------------------
    assets_df = _strip_blank(assets_df)
    if len(assets_df) == 0:
        raise DataValidationError(
            "No assets found in 'Assets' sheet",
            details={"sheet": "Assets"},
            suggestions=["Add at least one asset row to the Assets sheet"],
        )
    col_norm = _normalise_cols(assets_df)

    name_col = _get_col(assets_df, col_norm, "name", sheet="Assets")
    ret_col = _get_col(assets_df, col_norm, "expected_return", sheet="Assets")
    sector_col = _get_col(assets_df, col_norm, "sector", sheet="Assets", required=False)

    assets: list[Asset] = []
    asset_names: list[str] = []
    for row_idx, row in assets_df.iterrows():
        name = _clean_string(row[name_col])
        if not name:
            continue
        try:
            ret = _parse_number(row[ret_col], context=f"Assets row {row_idx + 1}")
        except ValueError as exc:
            raise DataValidationError(
                str(exc),
                details={"sheet": "Assets", "row": row_idx + 1, "column": ret_col},
                suggestions=["Use decimal (0.08) or percentage (8%) format"],
            ) from exc
        sector = _clean_string(row[sector_col]) if sector_col else None
        assets.append(Asset(name=name, expected_return=ret, sector=sector or None))
        asset_names.append(name)

    if not assets:
        raise DataValidationError(
            "No assets found in 'Assets' sheet",
            details={"sheet": "Assets"},
            suggestions=["Add at least one asset row to the Assets sheet"],
        )

    # ---- Parse covariance matrix --------------------------------------------
    cov_df = _strip_blank(cov_df)
    n = len(assets)
    # First column is asset name labels; remaining columns are values
    value_cols = cov_df.columns[1:]
    if len(value_cols) < n:
        raise DataValidationError(
            f"Covariance sheet has {len(value_cols)} value columns but {n} assets were found",
            details={"sheet": "Covariance", "expected_cols": n, "found_cols": len(value_cols)},
            suggestions=["Ensure the Covariance sheet has one row and one column per asset"],
        )
    if len(cov_df) < n:
        raise DataValidationError(
            f"Covariance sheet has {len(cov_df)} data rows but {n} assets were found",
            details={"sheet": "Covariance", "expected_rows": n, "found_rows": len(cov_df)},
        )

    cov: list[list[float]] = []
    for i in range(n):
        row_vals = []
        for j, col in enumerate(value_cols[:n]):
            raw = cov_df.iloc[i][col]
            try:
                row_vals.append(_parse_number(raw, context=f"Covariance[{i},{j}]"))
            except ValueError as exc:
                raise DataValidationError(
                    str(exc),
                    details={"sheet": "Covariance", "row": i + 2, "column": str(col)},
                ) from exc
        cov.append(row_vals)

    # ---- Parse constraints ---------------------------------------------------
    pc_kwargs: dict = {}
    if constraints_df is not None:
        constraints_df = _strip_blank(constraints_df)
        c_norm = _normalise_cols(constraints_df)
        param_col = _get_col(constraints_df, c_norm, "parameter", sheet="Constraints")
        val_col = _get_col(constraints_df, c_norm, "value", sheet="Constraints")

        for _, row in constraints_df.iterrows():
            param = _clean_string(row[param_col]).lower().replace(" ", "_")
            raw_val = row[val_col]
            if not param:
                continue
            try:
                val = _parse_number(raw_val, context=f"Constraints/{param}")
            except ValueError:
                continue  # skip unparseable
            pc_kwargs[param] = val

    risk_aversion = pc_kwargs.pop("risk_aversion", 1.0)
    pc = PortfolioConstraints(**pc_kwargs) if pc_kwargs else PortfolioConstraints()

    return PortfolioModel(
        assets=assets,
        covariance_matrix=cov,
        risk_aversion=risk_aversion,
        constraints=pc,
    )


# ---------------------------------------------------------------------------
# Scheduling parser
# ---------------------------------------------------------------------------


def _parse_scheduling(dfs: dict[str, pd.DataFrame]) -> SchedulingModel:
    sheets_norm = {k.strip().lower(): k for k in dfs}

    def _get_sheet(key: str) -> pd.DataFrame:
        actual = sheets_norm.get(key)
        if actual is None:
            raise DataValidationError(
                f"Required sheet '{key}' not found",
                details={"found_sheets": list(dfs.keys()), "required": key},
                suggestions=[f"Add a sheet named '{key}' to the workbook"],
            )
        return dfs[actual]

    workers_df = _strip_blank(_get_sheet("workers"))
    shifts_df = _strip_blank(_get_sheet("shifts"))
    constraints_df = _strip_blank(_get_sheet("constraints")) if "constraints" in sheets_norm else None

    # ---- Workers -------------------------------------------------------------
    w_norm = _normalise_cols(workers_df)
    name_col = _get_col(workers_df, w_norm, "name", sheet="Workers")
    hours_col = _get_col(workers_df, w_norm, "max_hours", sheet="Workers")
    skills_col = _get_col(workers_df, w_norm, "skills", sheet="Workers", required=False)
    unavail_col = _get_col(workers_df, w_norm, "unavailable_shifts", sheet="Workers", required=False)

    workers: list[Worker] = []
    for row_idx, row in workers_df.iterrows():
        name = _clean_string(row[name_col])
        if not name:
            continue
        try:
            max_hours = _parse_number(row[hours_col], context=f"Workers/{name}/max_hours")
        except ValueError as exc:
            raise DataValidationError(
                str(exc),
                details={"sheet": "Workers", "row": row_idx + 1, "column": hours_col},
            ) from exc
        skills = _parse_str_list(row[skills_col] if skills_col else None)
        unavail = _parse_str_list(row[unavail_col] if unavail_col else None)
        workers.append(Worker(
            name=name,
            max_hours=max_hours,
            skills=skills or None,
            unavailable_shifts=unavail or None,
        ))

    if not workers:
        raise DataValidationError(
            "No workers found in 'Workers' sheet",
            details={"sheet": "Workers"},
        )

    # ---- Shifts --------------------------------------------------------------
    s_norm = _normalise_cols(shifts_df)
    sname_col = _get_col(shifts_df, s_norm, "name", sheet="Shifts")
    dur_col = _get_col(shifts_df, s_norm, "duration_hours", sheet="Shifts")
    req_col = _get_col(shifts_df, s_norm, "required_workers", sheet="Shifts")
    req_skills_col = _get_col(shifts_df, s_norm, "required_skills", sheet="Shifts", required=False)

    shifts: list[Shift] = []
    for row_idx, row in shifts_df.iterrows():
        name = _clean_string(row[sname_col])
        if not name:
            continue
        try:
            duration = _parse_number(row[dur_col], context=f"Shifts/{name}/duration_hours")
            required = int(_parse_number(row[req_col], context=f"Shifts/{name}/required_workers"))
        except ValueError as exc:
            raise DataValidationError(
                str(exc),
                details={"sheet": "Shifts", "row": row_idx + 1},
            ) from exc
        req_skills = _parse_str_list(row[req_skills_col] if req_skills_col else None)
        shifts.append(Shift(
            name=name,
            duration_hours=duration,
            required_workers=required,
            required_skills=req_skills or None,
        ))

    if not shifts:
        raise DataValidationError(
            "No shifts found in 'Shifts' sheet",
            details={"sheet": "Shifts"},
        )

    # ---- Constraints ---------------------------------------------------------
    horizon = 7
    max_consec = 5
    if constraints_df is not None:
        c_norm = _normalise_cols(constraints_df)
        param_col = _get_col(constraints_df, c_norm, "parameter", sheet="Constraints")
        val_col = _get_col(constraints_df, c_norm, "value", sheet="Constraints")
        for _, row in constraints_df.iterrows():
            param = _clean_string(row[param_col]).lower().replace(" ", "_")
            raw_val = row[val_col]
            try:
                val = int(_parse_number(raw_val))
            except (ValueError, TypeError):
                continue
            if param == "planning_horizon_days":
                horizon = val
            elif param == "max_consecutive_days":
                max_consec = val

    return SchedulingModel(
        workers=workers,
        shifts=shifts,
        planning_horizon_days=horizon,
        max_consecutive_days=max_consec,
    )


# ---------------------------------------------------------------------------
# Transport parser (returns generic LPModel)
# ---------------------------------------------------------------------------


def _parse_transport_lp(dfs: dict[str, pd.DataFrame]) -> LPModel:
    """Parse a transportation problem template into a generic LPModel."""
    sheets_norm = {k.strip().lower(): k for k in dfs}

    def _get_sheet(key: str) -> pd.DataFrame:
        actual = sheets_norm.get(key)
        if actual is None:
            raise DataValidationError(
                f"Required sheet '{key}' not found",
                details={"found_sheets": list(dfs.keys()), "required": key},
                suggestions=[f"Add a '{key}' sheet to the workbook"],
            )
        return dfs[actual]

    origins_df = _strip_blank(_get_sheet("origins"))
    dests_df = _strip_blank(_get_sheet("destinations"))
    costs_df = _strip_blank(_get_sheet("costs"))

    # Parse origins
    o_norm = _normalise_cols(origins_df)
    o_name_col = _get_col(origins_df, o_norm, "name", sheet="Origins")
    o_supply_col = _get_col(origins_df, o_norm, "supply", sheet="Origins")

    origins: list[tuple[str, float]] = []
    for _, row in origins_df.iterrows():
        name = _clean_string(row[o_name_col])
        if not name:
            continue
        supply = _parse_number(row[o_supply_col], context="Origins/supply")
        origins.append((name, supply))

    # Parse destinations
    d_norm = _normalise_cols(dests_df)
    d_name_col = _get_col(dests_df, d_norm, "name", sheet="Destinations")
    d_demand_col = _get_col(dests_df, d_norm, "demand", sheet="Destinations")

    destinations: list[tuple[str, float]] = []
    for _, row in dests_df.iterrows():
        name = _clean_string(row[d_name_col])
        if not name:
            continue
        demand = _parse_number(row[d_demand_col], context="Destinations/demand")
        destinations.append((name, demand))

    if not origins or not destinations:
        raise DataValidationError(
            "Origins or Destinations sheet is empty",
            details={"num_origins": len(origins), "num_destinations": len(destinations)},
        )

    # Parse cost matrix: first column is origin labels, remaining are values
    costs_df_clean = _strip_blank(costs_df)
    dest_names = [d[0] for d in destinations]
    origin_names = [o[0] for o in origins]
    n_orig = len(origins)
    n_dest = len(destinations)

    cost_matrix: list[list[float]] = []
    for i in range(min(n_orig, len(costs_df_clean))):
        row = costs_df_clean.iloc[i]
        cost_row = []
        for j in range(1, n_dest + 1):
            if j < len(row):
                cost_row.append(_parse_number(row.iloc[j], context=f"Costs[{i},{j}]"))
            else:
                raise DataValidationError(
                    f"Costs sheet missing column for destination {j}",
                    details={"sheet": "Costs", "row": i + 1},
                )
        cost_matrix.append(cost_row)

    # Build LP: variables x_{origin}_{dest}, minimize total cost
    variables: list[LPVariable] = []
    obj_coeffs: dict[str, float] = {}
    for i, (o_name, _) in enumerate(origins):
        for j, (d_name, _) in enumerate(destinations):
            vname = f"x_{o_name}_{d_name}"
            variables.append(LPVariable(name=vname, lower_bound=0.0))
            obj_coeffs[vname] = cost_matrix[i][j]

    # Supply constraints
    constraints: list[LinearConstraint] = []
    for i, (o_name, supply) in enumerate(origins):
        coeffs = {
            f"x_{o_name}_{d_name}": 1.0
            for d_name, _ in destinations
        }
        constraints.append(LinearConstraint(
            name=f"supply_{o_name}",
            coefficients=coeffs,
            sense="<=",
            rhs=supply,
        ))

    # Demand constraints
    for j, (d_name, demand) in enumerate(destinations):
        coeffs = {
            f"x_{o_name}_{d_name}": 1.0
            for o_name, _ in origins
        }
        constraints.append(LinearConstraint(
            name=f"demand_{d_name}",
            coefficients=coeffs,
            sense=">=",
            rhs=demand,
        ))

    return LPModel(
        name="transport",
        variables=variables,
        constraints=constraints,
        objective=LinearObjective(sense="minimize", coefficients=obj_coeffs),
    )


# ---------------------------------------------------------------------------
# Generic LP parser
# ---------------------------------------------------------------------------


def _parse_generic_lp(dfs: dict[str, pd.DataFrame]) -> MIPModel:
    """Parse generic LP/MIP template DataFrames into a MIPModel."""
    sheets_norm = {k.strip().lower(): k for k in dfs}

    def _get_sheet(key: str) -> pd.DataFrame:
        actual = sheets_norm.get(key)
        if actual is None:
            raise DataValidationError(
                f"Required sheet '{key}' not found",
                details={"found_sheets": list(dfs.keys()), "required": key},
                suggestions=[f"Add a '{key}' sheet to the workbook"],
            )
        return dfs[actual]

    vars_df = _strip_blank(_get_sheet("variables"))
    constr_df = _strip_blank(_get_sheet("constraints"))
    obj_df = _strip_blank(_get_sheet("objective"))

    # ---- Variables -----------------------------------------------------------
    v_norm = _normalise_cols(vars_df)
    vname_col = _get_col(vars_df, v_norm, "name", sheet="Variables")
    lb_col = _get_col(vars_df, v_norm, "lower_bound", sheet="Variables", required=False)
    ub_col = _get_col(vars_df, v_norm, "upper_bound", sheet="Variables", required=False)
    vtype_col = _get_col(vars_df, v_norm, "type", sheet="Variables", required=False)

    variables: list[MIPVariable] = []
    for row_idx, row in vars_df.iterrows():
        name = _clean_string(row[vname_col])
        if not name:
            continue
        lb = _parse_optional_number(row[lb_col] if lb_col else None, default=0.0)
        ub = _parse_optional_number(row[ub_col] if ub_col else None, default=None)
        raw_type = _clean_string(row[vtype_col]).lower() if vtype_col else "continuous"
        if raw_type not in ("continuous", "integer", "binary"):
            raw_type = "continuous"
        variables.append(MIPVariable(name=name, lower_bound=lb, upper_bound=ub, var_type=raw_type))

    if not variables:
        raise DataValidationError(
            "No variables found in 'Variables' sheet",
            details={"sheet": "Variables"},
        )

    var_names = {v.name for v in variables}

    # ---- Constraints ---------------------------------------------------------
    c_norm = _normalise_cols(constr_df)
    cname_col = _get_col(constr_df, c_norm, "name", sheet="Constraints")
    coeff_col = _get_col(constr_df, c_norm, "coefficients", sheet="Constraints")
    sense_col = _get_col(constr_df, c_norm, "sense", sheet="Constraints")
    rhs_col = _get_col(constr_df, c_norm, "rhs", sheet="Constraints")

    constraints: list[LinearConstraint] = []
    for row_idx, row in constr_df.iterrows():
        cname = _clean_string(row[cname_col])
        if not cname:
            continue
        coeff_str = _clean_string(row[coeff_col])
        sense = _clean_string(row[sense_col])
        try:
            rhs = _parse_number(row[rhs_col], context=f"Constraints/{cname}/rhs")
        except ValueError as exc:
            raise DataValidationError(
                str(exc),
                details={"sheet": "Constraints", "row": row_idx + 1, "column": rhs_col},
            ) from exc

        # Parse coefficients dict (JSON-like string)
        try:
            import json
            coeffs_raw = json.loads(coeff_str)
            coeffs = {str(k): float(v) for k, v in coeffs_raw.items()}
        except Exception as exc:
            raise DataValidationError(
                f"Cannot parse coefficients in constraint '{cname}': {coeff_str!r}",
                details={"sheet": "Constraints", "row": row_idx + 1, "column": coeff_col},
                suggestions=['Use JSON format: {"x1": 2, "x2": 1}'],
            ) from exc

        if sense not in ("<=", ">=", "=="):
            raise DataValidationError(
                f"Invalid constraint sense '{sense}' in constraint '{cname}'",
                details={"sheet": "Constraints", "row": row_idx + 1, "column": sense_col},
                suggestions=["Use '<=', '>=', or '=='"],
            )
        constraints.append(LinearConstraint(name=cname, coefficients=coeffs, sense=sense, rhs=rhs))

    # ---- Objective -----------------------------------------------------------
    o_norm = _normalise_cols(obj_df)
    obj_sense_col = _get_col(obj_df, o_norm, "sense", sheet="Objective")
    obj_coeff_col = _get_col(obj_df, o_norm, "coefficients", sheet="Objective")

    obj_df_clean = _strip_blank(obj_df)
    if len(obj_df_clean) == 0:
        raise DataValidationError(
            "No objective found in 'Objective' sheet",
            details={"sheet": "Objective"},
        )

    obj_row = obj_df_clean.iloc[0]
    sense = _clean_string(obj_row[obj_sense_col]).lower()
    if sense not in ("minimize", "maximise", "maximize", "minimise"):
        raise DataValidationError(
            f"Invalid objective sense: '{sense}'",
            details={"sheet": "Objective", "found": sense},
            suggestions=["Use 'minimize' or 'maximize'"],
        )
    sense = "minimize" if "minim" in sense else "maximize"

    try:
        import json
        obj_coeffs_raw = json.loads(_clean_string(obj_row[obj_coeff_col]))
        obj_coeffs = {str(k): float(v) for k, v in obj_coeffs_raw.items()}
    except Exception as exc:
        raise DataValidationError(
            f"Cannot parse objective coefficients: {obj_row[obj_coeff_col]!r}",
            details={"sheet": "Objective"},
            suggestions=['Use JSON format: {"x1": 5, "x2": 4}'],
        ) from exc

    return MIPModel(
        name="generic_lp",
        variables=variables,
        constraints=constraints,
        objective=LinearObjective(sense=sense, coefficients=obj_coeffs),
    )
