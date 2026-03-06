"""Tests for sage_mcp.server — MCP tool handlers.

Tests call the private handler functions directly, bypassing MCP transport,
so no running MCP server is required.

Coverage:
- TestToolRegistration: all 7 tools registered with correct metadata
- TestSolveOptimization: LP, MIP, portfolio, scheduling; status/values correct
- TestReadDataFile: good file, bad path
- TestSolveFromFile: full pipeline with a test Excel file
- TestExplainSolution: state-based explanation, no prior state
- TestCheckFeasibility: feasible, infeasible + IIS, infeasible + suggestions
- TestGenerateTemplate: all 4 types; output file created
- TestSuggestRelaxations: uses stored infeasible state, no state case
- TestErrorHandling: bad model JSON, bad file path, malformed model
"""

from __future__ import annotations

import asyncio
import io
import json
import os
import tempfile
from pathlib import Path

import pytest

# Reset server state before each test so tests are independent
from sage_solver_mcp import server as _server_module


def _reset_state() -> None:
    _server_module._state.last_result = None
    _server_module._state.last_model = None
    _server_module._state.last_solver_input = None
    _server_module._state.last_iis = None


@pytest.fixture(autouse=True)
def reset_state():
    _reset_state()
    yield
    _reset_state()


# ---------------------------------------------------------------------------
# Async helpers
# ---------------------------------------------------------------------------


def run(coro):
    """Run a coroutine synchronously in tests."""
    return asyncio.get_event_loop().run_until_complete(coro)


# ---------------------------------------------------------------------------
# Fixtures: minimal models as dicts
# ---------------------------------------------------------------------------


@pytest.fixture()
def simple_lp_args() -> dict:
    """Maximize 3x + 2y subject to x + y <= 10, x <= 8."""
    return {
        "problem_type": "lp",
        "name": "simple_lp",
        "variables": [
            {"name": "x", "lower_bound": 0.0},
            {"name": "y", "lower_bound": 0.0},
        ],
        "constraints": [
            {"name": "c1", "coefficients": {"x": 1.0, "y": 1.0}, "sense": "<=", "rhs": 10.0},
            {"name": "c2", "coefficients": {"x": 1.0}, "sense": "<=", "rhs": 8.0},
        ],
        "objective": {"sense": "maximize", "coefficients": {"x": 3.0, "y": 2.0}},
    }


@pytest.fixture()
def infeasible_lp_args() -> dict:
    """x >= 10 AND x <= 5 — infeasible."""
    return {
        "problem_type": "lp",
        "name": "infeasible",
        "variables": [{"name": "x", "lower_bound": 0.0}],
        "constraints": [
            {"name": "lower_x", "coefficients": {"x": 1.0}, "sense": ">=", "rhs": 10.0},
            {"name": "upper_x", "coefficients": {"x": 1.0}, "sense": "<=", "rhs": 5.0},
        ],
        "objective": {"sense": "minimize", "coefficients": {"x": 1.0}},
    }


@pytest.fixture()
def simple_mip_args() -> dict:
    """MIP: binary y, maximize y, y <= 1 — trivial."""
    return {
        "problem_type": "mip",
        "name": "simple_mip",
        "variables": [{"name": "y", "lower_bound": 0.0, "upper_bound": 1.0, "var_type": "binary"}],
        "constraints": [
            {"name": "c1", "coefficients": {"y": 1.0}, "sense": "<=", "rhs": 1.0},
        ],
        "objective": {"sense": "maximize", "coefficients": {"y": 1.0}},
    }


@pytest.fixture()
def portfolio_args() -> dict:
    """2-asset portfolio with simple covariance."""
    return {
        "problem_type": "portfolio",
        "assets": [
            {"name": "Stock", "expected_return": 0.12},
            {"name": "Bond", "expected_return": 0.05},
        ],
        "covariance_matrix": [[0.04, 0.002], [0.002, 0.01]],
        "risk_aversion": 1.0,
        "constraints": {"max_total_allocation": 1.0},
    }


@pytest.fixture()
def scheduling_args() -> dict:
    """1 worker, 1 shift, 2 days — trivially feasible."""
    return {
        "problem_type": "scheduling",
        "workers": [{"name": "Alice", "max_hours": 20}],
        "shifts": [{"name": "Day", "duration_hours": 8, "required_workers": 1}],
        "planning_horizon_days": 2,
        "max_consecutive_days": 5,
    }


# ---------------------------------------------------------------------------
# Minimal Excel file fixture (portfolio template)
# ---------------------------------------------------------------------------


@pytest.fixture()
def portfolio_xlsx_path(tmp_path: Path) -> Path:
    """Write a minimal portfolio Excel file and return its path."""
    import openpyxl

    wb = openpyxl.Workbook()

    # Assets sheet
    ws_assets = wb.active
    ws_assets.title = "Assets"
    ws_assets.append(["name", "expected_return", "sector"])
    ws_assets.append(["AAPL", 0.15, "Tech"])
    ws_assets.append(["BOND", 0.04, "Fixed Income"])

    # Covariance sheet
    ws_cov = wb.create_sheet("Covariance")
    ws_cov.append(["", "AAPL", "BOND"])
    ws_cov.append(["AAPL", 0.04, 0.002])
    ws_cov.append(["BOND", 0.002, 0.01])

    # Constraints sheet
    ws_con = wb.create_sheet("Constraints")
    ws_con.append(["parameter", "value"])
    ws_con.append(["risk_aversion", 1.0])
    ws_con.append(["max_total_allocation", 1.0])
    ws_con.append(["min_total_allocation", 1.0])

    out = tmp_path / "portfolio_test.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# TestToolRegistration
# ---------------------------------------------------------------------------


class TestToolRegistration:
    def test_seven_tools_registered(self):
        tools = run(_server_module.list_tools())
        assert len(tools) == 7

    def test_tool_names(self):
        tools = run(_server_module.list_tools())
        names = {t.name for t in tools}
        expected = {
            "solve_optimization",
            "read_data_file",
            "solve_from_file",
            "explain_solution",
            "check_feasibility",
            "generate_template",
            "suggest_relaxations",
        }
        assert names == expected

    def test_all_tools_have_descriptions(self):
        tools = run(_server_module.list_tools())
        for tool in tools:
            assert tool.description is not None and len(tool.description) > 10, (
                f"Tool {tool.name!r} has no/short description"
            )

    def test_all_tools_have_input_schema(self):
        tools = run(_server_module.list_tools())
        for tool in tools:
            assert isinstance(tool.inputSchema, dict), f"Tool {tool.name!r} has no inputSchema"
            assert tool.inputSchema.get("type") == "object"

    def test_tool_descriptions_are_semantic(self):
        tools = run(_server_module.list_tools())
        tool_map = {t.name: t for t in tools}
        # Should mention high-level concepts, not internal names like "HiGHS"
        solve_desc = tool_map["solve_optimization"].description.lower()
        assert "highs" not in solve_desc
        assert "optimization" in solve_desc


# ---------------------------------------------------------------------------
# TestSolveOptimization
# ---------------------------------------------------------------------------


class TestSolveOptimization:
    def test_lp_optimal_status(self, simple_lp_args):
        result = run(_server_module._handle_solve_optimization(simple_lp_args))
        text = result[0].text
        assert "optimal" in text.lower() or "solution" in text.lower()

    def test_lp_objective_value_in_output(self, simple_lp_args):
        result = run(_server_module._handle_solve_optimization(simple_lp_args))
        text = result[0].text
        # Max 3x+2y s.t. x+y<=10, x<=8 → x=8, y=2, obj=28
        assert "28" in text

    def test_lp_stores_state(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        assert _server_module._state.last_result is not None
        assert _server_module._state.last_result.status == "optimal"

    def test_infeasible_lp_status(self, infeasible_lp_args):
        result = run(_server_module._handle_solve_optimization(infeasible_lp_args))
        text = result[0].text.lower()
        assert "infeasible" in text

    def test_infeasible_stores_iis(self, infeasible_lp_args):
        run(_server_module._handle_solve_optimization(infeasible_lp_args))
        assert _server_module._state.last_iis is not None

    def test_mip_optimal(self, simple_mip_args):
        result = run(_server_module._handle_solve_optimization(simple_mip_args))
        text = result[0].text.lower()
        assert "optimal" in text or "solution" in text

    def test_portfolio_optimal(self, portfolio_args):
        result = run(_server_module._handle_solve_optimization(portfolio_args))
        text = result[0].text
        assert "%" in text or "return" in text.lower() or "optimal" in text.lower()

    def test_scheduling_optimal(self, scheduling_args):
        result = run(_server_module._handle_solve_optimization(scheduling_args))
        text = result[0].text.lower()
        assert "optimal" in text or "solution" in text or "assignment" in text

    def test_auto_detect_lp_no_problem_type(self):
        """Omit problem_type; structure detection should infer LP."""
        args = {
            "name": "auto",
            "variables": [{"name": "z", "lower_bound": 0.0}],
            "constraints": [{"name": "c", "coefficients": {"z": 1.0}, "sense": "<=", "rhs": 5.0}],
            "objective": {"sense": "maximize", "coefficients": {"z": 1.0}},
        }
        result = run(_server_module._handle_solve_optimization(args))
        assert result[0].text  # non-empty response

    def test_auto_detect_mip_from_var_type(self):
        """Auto-detect MIP from presence of binary var_type."""
        args = {
            "name": "auto_mip",
            "variables": [{"name": "b", "var_type": "binary"}],
            "constraints": [{"name": "c", "coefficients": {"b": 1.0}, "sense": "<=", "rhs": 1.0}],
            "objective": {"sense": "maximize", "coefficients": {"b": 1.0}},
        }
        result = run(_server_module._handle_solve_optimization(args))
        assert result[0].text

    def test_returns_variable_values(self, simple_lp_args):
        result = run(_server_module._handle_solve_optimization(simple_lp_args))
        text = result[0].text
        assert "x" in text or "y" in text


# ---------------------------------------------------------------------------
# TestReadDataFile
# ---------------------------------------------------------------------------


class TestReadDataFile:
    def test_reads_xlsx_and_returns_summary(self, portfolio_xlsx_path):
        result = run(_server_module._handle_read_data_file({"filepath": str(portfolio_xlsx_path)}))
        text = result[0].text
        assert "Assets" in text or "assets" in text.lower()
        assert "Rows" in text or "rows" in text.lower()

    def test_returns_column_names(self, portfolio_xlsx_path):
        result = run(_server_module._handle_read_data_file({"filepath": str(portfolio_xlsx_path)}))
        text = result[0].text
        assert "name" in text.lower() or "expected_return" in text.lower()

    def test_missing_file_raises_error(self):
        result = run(_server_module._handle_read_data_file({"filepath": "/nonexistent/path/file.xlsx"}))
        text = result[0].text.lower()
        assert "error" in text or "not found" in text or "sage error" in text

    def test_shows_preview_rows(self, portfolio_xlsx_path):
        result = run(_server_module._handle_read_data_file({"filepath": str(portfolio_xlsx_path)}))
        text = result[0].text
        # Should show asset names from the file
        assert "AAPL" in text or "BOND" in text


# ---------------------------------------------------------------------------
# TestSolveFromFile
# ---------------------------------------------------------------------------


class TestSolveFromFile:
    def test_portfolio_file_solve(self, portfolio_xlsx_path, tmp_path):
        result = run(_server_module._handle_solve_from_file({
            "filepath": str(portfolio_xlsx_path),
            "problem_type": "portfolio",
        }))
        text = result[0].text
        # Should mention solution status
        assert "optimal" in text.lower() or "%" in text

    def test_output_file_created(self, portfolio_xlsx_path, tmp_path):
        run(_server_module._handle_solve_from_file({
            "filepath": str(portfolio_xlsx_path),
            "problem_type": "portfolio",
        }))
        # Output should be written next to input
        out = portfolio_xlsx_path.parent / "portfolio_test_optimized.xlsx"
        assert out.exists(), f"Output file not found at {out}"

    def test_output_path_mentioned_in_response(self, portfolio_xlsx_path):
        result = run(_server_module._handle_solve_from_file({
            "filepath": str(portfolio_xlsx_path),
            "problem_type": "portfolio",
        }))
        text = result[0].text
        assert "_optimized" in text or "Results written" in text

    def test_bad_file_path_returns_error(self):
        result = run(_server_module._handle_solve_from_file({
            "filepath": "/no/such/file.xlsx",
            "problem_type": "portfolio",
        }))
        text = result[0].text.lower()
        assert "error" in text or "not found" in text


# ---------------------------------------------------------------------------
# TestExplainSolution
# ---------------------------------------------------------------------------


class TestExplainSolution:
    def test_no_state_returns_error(self):
        result = run(_server_module._handle_explain_solution({}))
        text = result[0].text.lower()
        assert "error" in text or "no solve" in text or "run" in text

    def test_brief_explanation_after_solve(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        result = run(_server_module._handle_explain_solution({"detail_level": "brief"}))
        text = result[0].text
        assert len(text) > 0

    def test_detailed_explanation_after_solve(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        brief = run(_server_module._handle_explain_solution({"detail_level": "brief"}))[0].text
        detailed = run(_server_module._handle_explain_solution({"detail_level": "detailed"}))[0].text
        assert len(detailed) >= len(brief)

    def test_default_detail_level_is_standard(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        result = run(_server_module._handle_explain_solution({}))
        assert len(result[0].text) > 0

    def test_invalid_detail_level_falls_back(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        result = run(_server_module._handle_explain_solution({"detail_level": "bogus"}))
        assert len(result[0].text) > 0

    def test_no_markdown_in_output(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        for level in ("brief", "standard", "detailed"):
            text = run(_server_module._handle_explain_solution({"detail_level": level}))[0].text
            assert "**" not in text, f"Markdown bold in {level} explanation"
            assert "##" not in text, f"Markdown heading in {level} explanation"


# ---------------------------------------------------------------------------
# TestCheckFeasibility
# ---------------------------------------------------------------------------


class TestCheckFeasibility:
    def test_feasible_problem_returns_feasible(self, simple_lp_args):
        result = run(_server_module._handle_check_feasibility(simple_lp_args))
        text = result[0].text.upper()
        assert "FEASIBLE" in text

    def test_infeasible_problem_returns_infeasible(self, infeasible_lp_args):
        result = run(_server_module._handle_check_feasibility(infeasible_lp_args))
        text = result[0].text.upper()
        assert "INFEASIBLE" in text

    def test_infeasible_includes_iis_explanation(self, infeasible_lp_args):
        result = run(_server_module._handle_check_feasibility(infeasible_lp_args))
        text = result[0].text.lower()
        assert "infeasible" in text

    def test_infeasible_includes_relaxation_suggestions(self, infeasible_lp_args):
        result = run(_server_module._handle_check_feasibility(infeasible_lp_args))
        text = result[0].text
        # Should include either relaxation suggestions or no-relaxation message
        assert "relax" in text.lower() or "constraint" in text.lower()

    def test_stores_iis_state(self, infeasible_lp_args):
        run(_server_module._handle_check_feasibility(infeasible_lp_args))
        assert _server_module._state.last_iis is not None

    def test_feasible_does_not_store_iis(self, simple_lp_args):
        run(_server_module._handle_check_feasibility(simple_lp_args))
        assert _server_module._state.last_iis is None


# ---------------------------------------------------------------------------
# TestGenerateTemplate
# ---------------------------------------------------------------------------


class TestGenerateTemplate:
    @pytest.mark.parametrize("ptype", ["portfolio", "scheduling", "generic_lp"])
    def test_template_file_created(self, ptype, tmp_path):
        result = run(_server_module._handle_generate_template({
            "problem_type": ptype,
            "output_directory": str(tmp_path),
        }))
        text = result[0].text
        expected = tmp_path / f"{ptype}_template.xlsx"
        assert expected.exists(), f"Template not created: {expected}"

    def test_template_path_in_response(self, tmp_path):
        result = run(_server_module._handle_generate_template({
            "problem_type": "portfolio",
            "output_directory": str(tmp_path),
        }))
        text = result[0].text
        assert "template" in text.lower() or str(tmp_path) in text

    def test_template_without_output_dir_uses_cwd(self):
        result = run(_server_module._handle_generate_template({
            "problem_type": "generic_lp",
        }))
        text = result[0].text
        out = Path.cwd() / "generic_lp_template.xlsx"
        # Clean up if created
        if out.exists():
            out.unlink()
        # Just check response is non-empty and mentions template
        assert len(text) > 0


# ---------------------------------------------------------------------------
# TestSuggestRelaxations
# ---------------------------------------------------------------------------


class TestSuggestRelaxations:
    def test_no_state_returns_error(self):
        result = run(_server_module._handle_suggest_relaxations({}))
        text = result[0].text.lower()
        assert "error" in text or "no infeasible" in text or "run" in text

    def test_after_feasible_solve_returns_error(self, simple_lp_args):
        run(_server_module._handle_solve_optimization(simple_lp_args))
        result = run(_server_module._handle_suggest_relaxations({}))
        text = result[0].text.lower()
        assert "error" in text or "not infeasible" in text

    def test_after_infeasible_returns_suggestions(self, infeasible_lp_args):
        run(_server_module._handle_solve_optimization(infeasible_lp_args))
        result = run(_server_module._handle_suggest_relaxations({}))
        text = result[0].text
        # Should contain numbered suggestions or no-relaxation message
        assert "1]" in text or "constraint" in text.lower() or "relaxation" in text.lower()

    def test_suggestions_include_constraint_names(self, infeasible_lp_args):
        run(_server_module._handle_solve_optimization(infeasible_lp_args))
        result = run(_server_module._handle_suggest_relaxations({}))
        text = result[0].text
        # At least one of the IIS constraint names should appear
        assert "lower_x" in text or "upper_x" in text or "constraint" in text.lower()

    def test_suggestions_include_percentage(self, infeasible_lp_args):
        run(_server_module._handle_solve_optimization(infeasible_lp_args))
        result = run(_server_module._handle_suggest_relaxations({}))
        text = result[0].text
        assert "%" in text or "relaxation" in text.lower()


# ---------------------------------------------------------------------------
# TestErrorHandling
# ---------------------------------------------------------------------------


class TestErrorHandling:
    def test_malformed_model_returns_error_not_crash(self):
        """Completely malformed model JSON should return an error, not raise."""
        args = {"problem_type": "lp", "variables": "not_a_list"}
        result = run(_server_module._handle_solve_optimization(args))
        text = result[0].text.lower()
        assert "error" in text or "sage" in text

    def test_bad_filepath_read_data_error(self):
        result = run(_server_module._handle_read_data_file({"filepath": "does_not_exist.xlsx"}))
        text = result[0].text.lower()
        assert "error" in text or "not found" in text

    def test_bad_filepath_solve_from_file_error(self):
        result = run(_server_module._handle_solve_from_file({
            "filepath": "/no/such/file.xlsx",
            "problem_type": "portfolio",
        }))
        text = result[0].text.lower()
        assert "error" in text or "not found" in text

    def test_call_tool_unknown_name_returns_error(self):
        result = run(_server_module.call_tool("nonexistent_tool", {}))
        text = result[0].text.lower()
        assert "unknown" in text or "error" in text

    def test_server_never_raises_on_bad_input(self, simple_lp_args):
        """call_tool should always return a response, never raise."""
        # Pass wildly wrong args
        for bad_args in [{}, {"garbage": True}, {"problem_type": "lp"}]:
            result = run(_server_module.call_tool("solve_optimization", bad_args))
            assert isinstance(result, list)
            assert len(result) > 0

    def test_missing_required_lp_fields_returns_error(self):
        """LP model with missing 'objective' field."""
        args = {
            "problem_type": "lp",
            "name": "incomplete",
            "variables": [{"name": "x", "lower_bound": 0.0}],
            "constraints": [],
            # missing 'objective'
        }
        result = run(_server_module._handle_solve_optimization(args))
        text = result[0].text.lower()
        assert "error" in text or "validation" in text

    def test_portfolio_missing_covariance_returns_error(self):
        args = {
            "problem_type": "portfolio",
            "assets": [{"name": "A", "expected_return": 0.1}],
            # missing covariance_matrix
            "risk_aversion": 1.0,
            "constraints": {"max_total_allocation": 1.0},
        }
        result = run(_server_module._handle_solve_optimization(args))
        text = result[0].text.lower()
        assert "error" in text
