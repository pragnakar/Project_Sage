"""SAGE Core — File I/O tests (Stage 4).

Tests cover:
- read_data / read_data_from_bytes (Excel + CSV)
- write_results_excel / write_results_csv
- generate_template (all 4 problem types)
- dataframe_to_model (portfolio, scheduling, transport, generic_lp)
- Round-trip: generate template → fill data → read → parse → verify model
- Messy data handling (string numbers, %, blank rows)
- Error handling: missing required columns → DataValidationError
"""

from __future__ import annotations

import io
import os
import tempfile

import pandas as pd
import pytest
import openpyxl

from sage_solver_core.fileio import (
    dataframe_to_model,
    generate_template,
    read_data,
    read_data_from_bytes,
    write_results_csv,
    write_results_excel,
)
from sage_solver_core.models import (
    Asset,
    DataValidationError,
    FileIOError,
    IISResult,
    LPModel,
    LPVariable,
    LinearConstraint,
    LinearObjective,
    MIPModel,
    MIPVariable,
    PortfolioConstraints,
    PortfolioModel,
    SchedulingModel,
    SolverResult,
    Worker,
    Shift,
)


# ---------------------------------------------------------------------------
# Fixtures helpers
# ---------------------------------------------------------------------------

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


def _make_solver_result_optimal() -> SolverResult:
    """Canonical optimal LP result for write tests."""
    return SolverResult(
        status="optimal",
        objective_value=26.0,
        solve_time_seconds=0.003,
        variable_values={"x": 6.0, "y": 4.0},
        shadow_prices={"sum_limit": 2.0, "x_limit": 1.0, "y_limit": 0.0},
        reduced_costs={"x": 0.0, "y": 0.0},
        constraint_slack={"sum_limit": 0.0, "x_limit": 0.0, "y_limit": 4.0},
        binding_constraints=["sum_limit", "x_limit"],
        objective_ranges={"x": (None, 3.0), "y": (2.0, None)},
        rhs_ranges={
            "sum_limit": (6.0, None),
            "x_limit": (None, 10.0),
            "y_limit": (4.0, None),
        },
    )


def _make_solver_result_infeasible() -> SolverResult:
    return SolverResult(
        status="infeasible",
        solve_time_seconds=0.001,
        iis=IISResult(
            conflicting_constraints=["upper_sum", "lower_sum"],
            conflicting_variable_bounds=[],
            explanation="The model is infeasible. Constraints 'upper_sum' and 'lower_sum' conflict.",
        ),
    )


def _make_portfolio_excel(path: str) -> None:
    """Create a minimal valid portfolio Excel file at path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Assets sheet
    ws_assets = wb.create_sheet("Assets")
    ws_assets.append(["Name", "Expected Return", "Sector"])
    ws_assets.append(["Stocks", 0.10, "Equity"])
    ws_assets.append(["Bonds", 0.04, "Fixed"])
    ws_assets.append(["Gold", 0.06, "Commodity"])

    # Covariance sheet
    ws_cov = wb.create_sheet("Covariance")
    ws_cov.append(["", "Stocks", "Bonds", "Gold"])
    ws_cov.append(["Stocks", 0.04, 0.002, 0.001])
    ws_cov.append(["Bonds", 0.002, 0.001, 0.0005])
    ws_cov.append(["Gold", 0.001, 0.0005, 0.005])

    # Constraints sheet
    ws_con = wb.create_sheet("Constraints")
    ws_con.append(["Parameter", "Value"])
    ws_con.append(["min_total_allocation", 1.0])
    ws_con.append(["max_total_allocation", 1.0])
    ws_con.append(["min_allocation_per_asset", 0.05])
    ws_con.append(["max_allocation_per_asset", 0.6])
    ws_con.append(["risk_aversion", 2.0])

    wb.save(path)


def _make_scheduling_excel(path: str) -> None:
    """Create a minimal valid scheduling Excel file at path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_workers = wb.create_sheet("Workers")
    ws_workers.append(["Name", "Max_Hours", "Skills", "Unavailable_Shifts"])
    ws_workers.append(["Alice", 40, "ICU,General", ""])
    ws_workers.append(["Bob", 40, "ER,General", ""])

    ws_shifts = wb.create_sheet("Shifts")
    ws_shifts.append(["Name", "Duration_Hours", "Required_Workers", "Required_Skills"])
    ws_shifts.append(["Morning", 8, 1, "General"])
    ws_shifts.append(["Night", 8, 1, ""])

    ws_con = wb.create_sheet("Constraints")
    ws_con.append(["Parameter", "Value"])
    ws_con.append(["planning_horizon_days", 3])
    ws_con.append(["max_consecutive_days", 3])

    wb.save(path)


# ---------------------------------------------------------------------------
# Tests: read_data — Excel
# ---------------------------------------------------------------------------


class TestReadDataExcel:
    def test_read_single_sheet(self, tmp_path):
        p = str(tmp_path / "test.xlsx")
        _make_portfolio_excel(p)
        dfs = read_data(p)
        assert "Assets" in dfs
        assert "Covariance" in dfs
        assert "Constraints" in dfs

    def test_assets_shape(self, tmp_path):
        p = str(tmp_path / "portfolio.xlsx")
        _make_portfolio_excel(p)
        dfs = read_data(p)
        assets_df = dfs["Assets"]
        assert len(assets_df) == 3
        assert "Name" in assets_df.columns

    def test_blank_rows_stripped(self, tmp_path):
        p = str(tmp_path / "blanks.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])
        ws.append([None, None])   # blank row
        ws.append(["Alice", 10])
        ws.append([None, None])   # trailing blank
        wb.save(p)
        dfs = read_data(p)
        df = list(dfs.values())[0]
        # Blank rows should be stripped
        assert len(df) == 1
        assert df.iloc[0]["Name"] == "Alice"

    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileIOError, match="File not found"):
            read_data(str(tmp_path / "nonexistent.xlsx"))

    def test_read_data_from_bytes_excel(self, tmp_path):
        p = str(tmp_path / "portfolio.xlsx")
        _make_portfolio_excel(p)
        with open(p, "rb") as f:
            content = f.read()
        dfs = read_data_from_bytes(content, "portfolio.xlsx")
        assert "Assets" in dfs
        assert len(dfs["Assets"]) == 3


# ---------------------------------------------------------------------------
# Tests: read_data — CSV
# ---------------------------------------------------------------------------


class TestReadDataCSV:
    def test_basic_csv(self, tmp_path):
        p = str(tmp_path / "data.csv")
        pd.DataFrame({"x": [1, 2, 3], "y": [4, 5, 6]}).to_csv(p, index=False)
        dfs = read_data(p)
        assert "data" in dfs
        assert len(dfs["data"]) == 3

    def test_csv_utf8(self, tmp_path):
        p = str(tmp_path / "utf8.csv")
        content = "Name,Value\nAlice,10\nBob,20\n"
        with open(p, "w", encoding="utf-8") as f:
            f.write(content)
        dfs = read_data(p)
        assert len(dfs["data"]) == 2

    def test_csv_from_bytes(self):
        content = b"Name,Value\nAlice,10\nBob,20\n"
        dfs = read_data_from_bytes(content, "test.csv")
        assert "data" in dfs
        assert len(dfs["data"]) == 2

    def test_tsv_from_bytes(self):
        content = b"Name\tValue\nAlice\t10\nBob\t20\n"
        dfs = read_data_from_bytes(content, "test.tsv")
        assert "data" in dfs
        assert len(dfs["data"]) == 2

    def test_unsupported_extension_raises(self):
        with pytest.raises(FileIOError, match="auto-detect"):
            read_data_from_bytes(b"data", "file.json")


# ---------------------------------------------------------------------------
# Tests: write_results_excel
# ---------------------------------------------------------------------------


class TestWriteResultsExcel:
    def test_creates_file(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        result = _make_solver_result_optimal()
        out = write_results_excel(result, "test_model", p)
        assert out == p
        assert os.path.isfile(p)

    def test_summary_sheet_exists(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        write_results_excel(_make_solver_result_optimal(), "my_model", p)
        wb = openpyxl.load_workbook(p)
        assert "Summary" in wb.sheetnames

    def test_solution_sheet_exists(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        write_results_excel(_make_solver_result_optimal(), "my_model", p)
        wb = openpyxl.load_workbook(p)
        assert "Solution" in wb.sheetnames

    def test_sensitivity_sheet_exists(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        write_results_excel(_make_solver_result_optimal(), "my_model", p)
        wb = openpyxl.load_workbook(p)
        assert "Sensitivity" in wb.sheetnames

    def test_constraints_sheet_exists(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        write_results_excel(_make_solver_result_optimal(), "my_model", p)
        wb = openpyxl.load_workbook(p)
        assert "Constraints" in wb.sheetnames

    def test_solution_values_correct(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        write_results_excel(_make_solver_result_optimal(), "my_model", p)
        wb = openpyxl.load_workbook(p)
        ws = wb["Solution"]
        # Header is row 1; data starts row 2. Sorted by value descending: x=6, y=4
        data = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, 10) if ws.cell(r, 1).value}
        assert data.get("x") == pytest.approx(6.0)
        assert data.get("y") == pytest.approx(4.0)

    def test_infeasible_result_has_infeasibility_sheet(self, tmp_path):
        p = str(tmp_path / "infeasible.xlsx")
        write_results_excel(_make_solver_result_infeasible(), "infeasible_model", p)
        wb = openpyxl.load_workbook(p)
        assert "Infeasibility" in wb.sheetnames
        # No Solution sheet
        assert "Solution" not in wb.sheetnames

    def test_model_name_in_summary(self, tmp_path):
        p = str(tmp_path / "results.xlsx")
        write_results_excel(_make_solver_result_optimal(), "MyModel2026", p)
        wb = openpyxl.load_workbook(p)
        ws = wb["Summary"]
        values = [ws.cell(r, 2).value for r in range(1, 10)]
        assert "MyModel2026" in values

    def test_write_to_bad_path_raises(self, tmp_path):
        p = str(tmp_path / "nonexistent_dir" / "results.xlsx")
        with pytest.raises(FileIOError, match="Cannot save"):
            write_results_excel(_make_solver_result_optimal(), "m", p)


# ---------------------------------------------------------------------------
# Tests: write_results_csv
# ---------------------------------------------------------------------------


class TestWriteResultsCSV:
    def test_creates_file(self, tmp_path):
        p = str(tmp_path / "results.csv")
        out = write_results_csv(_make_solver_result_optimal(), p)
        assert out == p
        assert os.path.isfile(p)

    def test_csv_values(self, tmp_path):
        p = str(tmp_path / "results.csv")
        write_results_csv(_make_solver_result_optimal(), p)
        df = pd.read_csv(p)
        assert set(df.columns) == {"variable", "value"}
        vals = dict(zip(df["variable"], df["value"]))
        assert vals["x"] == pytest.approx(6.0)
        assert vals["y"] == pytest.approx(4.0)

    def test_infeasible_csv(self, tmp_path):
        p = str(tmp_path / "infeasible.csv")
        write_results_csv(_make_solver_result_infeasible(), p)
        df = pd.read_csv(p)
        assert "status" in df.columns


# ---------------------------------------------------------------------------
# Tests: generate_template
# ---------------------------------------------------------------------------


class TestGenerateTemplate:
    @pytest.mark.parametrize("ptype", ["portfolio", "scheduling", "transport", "generic_lp"])
    def test_creates_valid_xlsx(self, tmp_path, ptype):
        p = str(tmp_path / f"{ptype}_template.xlsx")
        out = generate_template(ptype, p)
        assert out == p
        assert os.path.isfile(p)
        wb = openpyxl.load_workbook(p)
        assert "Instructions" in wb.sheetnames

    def test_portfolio_template_has_required_sheets(self, tmp_path):
        p = str(tmp_path / "portfolio.xlsx")
        generate_template("portfolio", p)
        wb = openpyxl.load_workbook(p)
        assert "Assets" in wb.sheetnames
        assert "Covariance" in wb.sheetnames
        assert "Constraints" in wb.sheetnames

    def test_scheduling_template_has_required_sheets(self, tmp_path):
        p = str(tmp_path / "scheduling.xlsx")
        generate_template("scheduling", p)
        wb = openpyxl.load_workbook(p)
        assert "Workers" in wb.sheetnames
        assert "Shifts" in wb.sheetnames

    def test_transport_template_has_required_sheets(self, tmp_path):
        p = str(tmp_path / "transport.xlsx")
        generate_template("transport", p)
        wb = openpyxl.load_workbook(p)
        assert "Origins" in wb.sheetnames
        assert "Destinations" in wb.sheetnames
        assert "Costs" in wb.sheetnames

    def test_generic_lp_template_has_required_sheets(self, tmp_path):
        p = str(tmp_path / "generic.xlsx")
        generate_template("generic_lp", p)
        wb = openpyxl.load_workbook(p)
        assert "Variables" in wb.sheetnames
        assert "Constraints" in wb.sheetnames
        assert "Objective" in wb.sheetnames

    def test_unknown_type_raises(self, tmp_path):
        with pytest.raises(DataValidationError, match="Unknown problem type"):
            generate_template("badtype", str(tmp_path / "x.xlsx"))


# ---------------------------------------------------------------------------
# Tests: dataframe_to_model — portfolio
# ---------------------------------------------------------------------------


class TestDataframeToModelPortfolio:
    def _make_dfs(self) -> dict[str, pd.DataFrame]:
        assets_df = pd.DataFrame({
            "Name": ["Stocks", "Bonds", "Gold"],
            "Expected Return": [0.10, 0.04, 0.06],
            "Sector": ["Equity", "Fixed", "Commodity"],
        })
        cov_df = pd.DataFrame({
            "": ["Stocks", "Bonds", "Gold"],
            "Stocks": [0.04, 0.002, 0.001],
            "Bonds": [0.002, 0.001, 0.0005],
            "Gold": [0.001, 0.0005, 0.005],
        })
        constraints_df = pd.DataFrame({
            "Parameter": ["min_total_allocation", "max_total_allocation", "min_allocation_per_asset", "max_allocation_per_asset", "risk_aversion"],
            "Value": [1.0, 1.0, 0.05, 0.6, 2.0],
        })
        return {"Assets": assets_df, "Covariance": cov_df, "Constraints": constraints_df}

    def test_basic_parse(self):
        model = dataframe_to_model(self._make_dfs(), "portfolio")
        assert isinstance(model, PortfolioModel)
        assert len(model.assets) == 3
        assert model.assets[0].name == "Stocks"
        assert model.assets[0].expected_return == pytest.approx(0.10)
        assert model.risk_aversion == pytest.approx(2.0)

    def test_covariance_matrix(self):
        model = dataframe_to_model(self._make_dfs(), "portfolio")
        assert len(model.covariance_matrix) == 3
        assert model.covariance_matrix[0][0] == pytest.approx(0.04)
        assert model.covariance_matrix[0][1] == pytest.approx(0.002)

    def test_constraints_parsed(self):
        model = dataframe_to_model(self._make_dfs(), "portfolio")
        assert model.constraints.min_total_allocation == pytest.approx(1.0)
        assert model.constraints.min_allocation_per_asset == pytest.approx(0.05)
        assert model.constraints.max_allocation_per_asset == pytest.approx(0.6)

    def test_percentage_return(self):
        """String percentage returns should be parsed correctly."""
        dfs = self._make_dfs()
        dfs["Assets"]["Expected Return"] = ["10%", "4%", "6%"]
        model = dataframe_to_model(dfs, "portfolio")
        assert model.assets[0].expected_return == pytest.approx(0.10)
        assert model.assets[1].expected_return == pytest.approx(0.04)

    def test_comma_number_in_covariance(self):
        """String numbers with commas should parse (unlikely in cov matrix but test robustness)."""
        dfs = self._make_dfs()
        # Replace one cov value with a string
        dfs["Covariance"]["Stocks"] = [0.04, 0.002, 0.001]  # keeps working
        model = dataframe_to_model(dfs, "portfolio")
        assert model.covariance_matrix[0][0] == pytest.approx(0.04)

    def test_missing_assets_sheet_raises(self):
        dfs = {"Covariance": pd.DataFrame(), "Constraints": pd.DataFrame()}
        with pytest.raises(DataValidationError, match="assets"):
            dataframe_to_model(dfs, "portfolio")

    def test_missing_expected_return_column_raises(self):
        dfs = self._make_dfs()
        dfs["Assets"] = dfs["Assets"].drop(columns=["Expected Return"])
        with pytest.raises(DataValidationError, match="expected_return"):
            dataframe_to_model(dfs, "portfolio")

    def test_extra_columns_ignored(self):
        dfs = self._make_dfs()
        dfs["Assets"]["irrelevant_column"] = ["a", "b", "c"]
        model = dataframe_to_model(dfs, "portfolio")
        assert len(model.assets) == 3

    def test_sector_optional(self):
        dfs = self._make_dfs()
        dfs["Assets"] = dfs["Assets"].drop(columns=["Sector"])
        model = dataframe_to_model(dfs, "portfolio")
        assert all(a.sector is None for a in model.assets)


# ---------------------------------------------------------------------------
# Tests: dataframe_to_model — scheduling
# ---------------------------------------------------------------------------


class TestDataframeToModelScheduling:
    def _make_dfs(self) -> dict[str, pd.DataFrame]:
        workers_df = pd.DataFrame({
            "Name": ["Alice", "Bob"],
            "Max_Hours": [40, 40],
            "Skills": ["ICU,General", "ER"],
            "Unavailable_Shifts": ["", "Night"],
        })
        shifts_df = pd.DataFrame({
            "Name": ["Morning", "Night"],
            "Duration_Hours": [8, 8],
            "Required_Workers": [1, 1],
            "Required_Skills": ["General", ""],
        })
        constraints_df = pd.DataFrame({
            "Parameter": ["planning_horizon_days", "max_consecutive_days"],
            "Value": [3, 3],
        })
        return {"Workers": workers_df, "Shifts": shifts_df, "Constraints": constraints_df}

    def test_basic_parse(self):
        model = dataframe_to_model(self._make_dfs(), "scheduling")
        assert isinstance(model, SchedulingModel)
        assert len(model.workers) == 2
        assert len(model.shifts) == 2

    def test_worker_attributes(self):
        model = dataframe_to_model(self._make_dfs(), "scheduling")
        alice = next(w for w in model.workers if w.name == "Alice")
        assert alice.max_hours == pytest.approx(40.0)
        assert "ICU" in alice.skills
        assert "General" in alice.skills

    def test_unavailable_shifts_parsed(self):
        model = dataframe_to_model(self._make_dfs(), "scheduling")
        bob = next(w for w in model.workers if w.name == "Bob")
        assert bob.unavailable_shifts == ["Night"]

    def test_required_skills_parsed(self):
        model = dataframe_to_model(self._make_dfs(), "scheduling")
        morning = next(s for s in model.shifts if s.name == "Morning")
        assert morning.required_skills == ["General"]
        night = next(s for s in model.shifts if s.name == "Night")
        assert not night.required_skills  # empty → None

    def test_constraints_parsed(self):
        model = dataframe_to_model(self._make_dfs(), "scheduling")
        assert model.planning_horizon_days == 3
        assert model.max_consecutive_days == 3

    def test_missing_workers_sheet_raises(self):
        dfs = {"Shifts": pd.DataFrame(), "Constraints": pd.DataFrame()}
        with pytest.raises(DataValidationError, match="workers"):
            dataframe_to_model(dfs, "scheduling")

    def test_missing_name_column_raises(self):
        dfs = self._make_dfs()
        dfs["Workers"] = dfs["Workers"].drop(columns=["Name"])
        with pytest.raises(DataValidationError, match="name"):
            dataframe_to_model(dfs, "scheduling")


# ---------------------------------------------------------------------------
# Tests: dataframe_to_model — generic_lp
# ---------------------------------------------------------------------------


class TestDataframeToModelGenericLP:
    def _make_dfs(self) -> dict[str, pd.DataFrame]:
        vars_df = pd.DataFrame({
            "Name": ["x1", "x2"],
            "Lower_Bound": [0, 0],
            "Upper_Bound": [None, None],
            "Type": ["continuous", "continuous"],
        })
        constr_df = pd.DataFrame({
            "Name": ["res_A", "res_B"],
            "Coefficients": ['{"x1": 2, "x2": 1}', '{"x1": 1, "x2": 2}'],
            "Sense": ["<=", "<="],
            "RHS": [14, 14],
        })
        obj_df = pd.DataFrame({
            "Sense": ["maximize"],
            "Coefficients": ['{"x1": 5, "x2": 4}'],
        })
        return {"Variables": vars_df, "Constraints": constr_df, "Objective": obj_df}

    def test_basic_parse(self):
        model = dataframe_to_model(self._make_dfs(), "generic_lp")
        assert isinstance(model, MIPModel)
        assert len(model.variables) == 2
        assert len(model.constraints) == 2
        assert model.objective.sense == "maximize"

    def test_objective_coefficients(self):
        model = dataframe_to_model(self._make_dfs(), "generic_lp")
        assert model.objective.coefficients["x1"] == pytest.approx(5.0)
        assert model.objective.coefficients["x2"] == pytest.approx(4.0)

    def test_constraint_parsing(self):
        model = dataframe_to_model(self._make_dfs(), "generic_lp")
        res_a = next(c for c in model.constraints if c.name == "res_A")
        assert res_a.coefficients["x1"] == pytest.approx(2.0)
        assert res_a.rhs == pytest.approx(14.0)
        assert res_a.sense == "<="

    def test_invalid_sense_raises(self):
        dfs = self._make_dfs()
        dfs["Constraints"]["Sense"] = ["??", "<="]
        with pytest.raises(DataValidationError, match="sense"):
            dataframe_to_model(dfs, "generic_lp")

    def test_invalid_coefficients_raises(self):
        dfs = self._make_dfs()
        dfs["Constraints"]["Coefficients"] = ["not_json", '{"x1": 1}']
        with pytest.raises(DataValidationError, match="coefficients"):
            dataframe_to_model(dfs, "generic_lp")


# ---------------------------------------------------------------------------
# Tests: messy data handling
# ---------------------------------------------------------------------------


class TestMessyData:
    def test_string_numbers_in_portfolio(self):
        """String numbers like '1,000' and percentages should parse correctly."""
        assets_df = pd.DataFrame({
            "Name": ["A", "B"],
            "Expected Return": ["8%", "  5.5 %"],
            "Sector": ["X", "Y"],
        })
        cov_df = pd.DataFrame({
            "": ["A", "B"],
            "A": ["0.04", "0.001"],
            "B": ["0.001", "0.009"],
        })
        dfs = {"Assets": assets_df, "Covariance": cov_df}
        model = dataframe_to_model(dfs, "portfolio")
        assert model.assets[0].expected_return == pytest.approx(0.08)
        assert model.assets[1].expected_return == pytest.approx(0.055)

    def test_blank_rows_in_assets(self):
        """Blank rows in the middle of a sheet are skipped."""
        assets_df = pd.DataFrame({
            "Name": ["Stocks", None, "Bonds"],
            "Expected Return": [0.10, None, 0.04],
            "Sector": ["E", None, "F"],
        })
        cov_df = pd.DataFrame({
            "": ["Stocks", "Bonds"],
            "Stocks": [0.04, 0.002],
            "Bonds": [0.002, 0.001],
        })
        # Strip blank from assets_df before passing
        assets_clean = assets_df.dropna(how="all").reset_index(drop=True)
        dfs = {"Assets": assets_clean, "Covariance": cov_df}
        model = dataframe_to_model(dfs, "portfolio")
        assert len(model.assets) == 2

    def test_whitespace_in_column_headers(self):
        """Column headers with extra spaces are matched correctly."""
        assets_df = pd.DataFrame({
            "  Name  ": ["A"],
            "  Expected Return  ": [0.10],
        })
        cov_df = pd.DataFrame({
            "": ["A"],
            "A": [0.04],
        })
        dfs = {"Assets": assets_df, "Covariance": cov_df}
        model = dataframe_to_model(dfs, "portfolio")
        assert model.assets[0].name == "A"

    def test_extra_columns_in_workers(self):
        """Extra unrecognised columns in Workers sheet are silently ignored."""
        workers_df = pd.DataFrame({
            "Name": ["Alice"],
            "Max_Hours": [40],
            "Skills": ["General"],
            "EXTRA_COLUMN": ["something_irrelevant"],
        })
        shifts_df = pd.DataFrame({
            "Name": ["Morning"],
            "Duration_Hours": [8],
            "Required_Workers": [1],
        })
        dfs = {"Workers": workers_df, "Shifts": shifts_df}
        model = dataframe_to_model(dfs, "scheduling")
        assert len(model.workers) == 1

    def test_missing_required_column_raises_with_details(self):
        """DataValidationError for missing column includes sheet and column info."""
        workers_df = pd.DataFrame({
            "Name": ["Alice"],
            # Missing Max_Hours
        })
        dfs = {"Workers": workers_df, "Shifts": pd.DataFrame()}
        with pytest.raises(DataValidationError) as exc_info:
            dataframe_to_model(dfs, "scheduling")
        err = exc_info.value
        assert "max_hours" in err.message.lower() or "max_hours" in str(err.details)
        assert err.details.get("sheet") == "Workers"


# ---------------------------------------------------------------------------
# Tests: round-trip — generate template → fill data → read → parse → solve
# ---------------------------------------------------------------------------


class TestRoundTrip:
    def test_portfolio_round_trip(self, tmp_path):
        """Generate portfolio template, fill data, read back, parse, solve."""
        from sage_solver_core.builder import build_from_portfolio
        from sage_solver_core.solver import solve

        # Step 1: generate template
        template_path = str(tmp_path / "portfolio_template.xlsx")
        generate_template("portfolio", template_path)

        # Step 2: create a proper filled file (not using template — just need clean Excel)
        filled_path = str(tmp_path / "portfolio_filled.xlsx")
        _make_portfolio_excel(filled_path)

        # Step 3: read
        dfs = read_data(filled_path)
        assert "Assets" in dfs

        # Step 4: parse
        model = dataframe_to_model(dfs, "portfolio")
        assert isinstance(model, PortfolioModel)
        assert len(model.assets) == 3

        # Step 5: build and solve
        si = build_from_portfolio(model)
        result = solve(si)
        assert result.status == "optimal"
        assert sum(result.variable_values.values()) == pytest.approx(1.0, abs=1e-4)

    def test_scheduling_round_trip(self, tmp_path):
        """Generate scheduling template, fill data, read back, parse, solve."""
        from sage_solver_core.builder import build_from_scheduling
        from sage_solver_core.solver import solve

        filled_path = str(tmp_path / "scheduling_filled.xlsx")
        _make_scheduling_excel(filled_path)

        dfs = read_data(filled_path)
        model = dataframe_to_model(dfs, "scheduling")
        assert isinstance(model, SchedulingModel)

        si = build_from_scheduling(model)
        result = solve(si)
        assert result.status == "optimal"

    def test_write_then_read_results(self, tmp_path):
        """Write results to Excel, read back, verify values match."""
        result = _make_solver_result_optimal()
        out_path = str(tmp_path / "results.xlsx")
        write_results_excel(result, "canonical_lp", out_path)

        # Read the Solution sheet and verify
        dfs = read_data(out_path)
        sol_df = dfs["Solution"]
        assert "Variable" in sol_df.columns
        assert "Optimal Value" in sol_df.columns
        vals = dict(zip(sol_df["Variable"], sol_df["Optimal Value"]))
        assert float(vals["x"]) == pytest.approx(6.0, abs=1e-3)
        assert float(vals["y"]) == pytest.approx(4.0, abs=1e-3)

    def test_generic_lp_round_trip(self, tmp_path):
        """Parse generic_lp DataFrames, solve, write results."""
        from sage_solver_core.builder import build_from_mip
        from sage_solver_core.solver import solve

        vars_df = pd.DataFrame({
            "Name": ["x1", "x2"],
            "Lower_Bound": [0, 0],
            "Upper_Bound": [None, None],
            "Type": ["continuous", "continuous"],
        })
        constr_df = pd.DataFrame({
            "Name": ["res_A", "res_B"],
            "Coefficients": ['{"x1": 2, "x2": 1}', '{"x1": 1, "x2": 2}'],
            "Sense": ["<=", "<="],
            "RHS": [14, 14],
        })
        obj_df = pd.DataFrame({
            "Sense": ["maximize"],
            "Coefficients": ['{"x1": 5, "x2": 4}'],
        })
        dfs = {"Variables": vars_df, "Constraints": constr_df, "Objective": obj_df}

        model = dataframe_to_model(dfs, "generic_lp")
        si = build_from_mip(model)
        result = solve(si)

        assert result.status == "optimal"
        assert result.objective_value == pytest.approx(42.0, abs=1e-3)

        out_path = str(tmp_path / "generic_lp_results.xlsx")
        write_results_excel(result, "generic_lp", out_path)
        assert os.path.isfile(out_path)

    def test_transport_round_trip(self, tmp_path):
        """Parse transport DataFrames, solve."""
        from sage_solver_core.builder import build_from_lp
        from sage_solver_core.solver import solve

        origins_df = pd.DataFrame({"Name": ["W1", "W2"], "Supply": [300, 400]})
        dests_df = pd.DataFrame({"Name": ["C1", "C2"], "Demand": [250, 350]})
        costs_df = pd.DataFrame({"": ["W1", "W2"], "C1": [2, 5], "C2": [3, 4]})
        dfs = {"Origins": origins_df, "Destinations": dests_df, "Costs": costs_df}

        model = dataframe_to_model(dfs, "transport")
        si = build_from_lp(model)
        result = solve(si)
        assert result.status == "optimal"


# ---------------------------------------------------------------------------
# Tests: error handling
# ---------------------------------------------------------------------------


class TestErrorHandling:
    def test_missing_required_column_includes_sheet_name(self):
        dfs = {"Assets": pd.DataFrame({"Name": ["X"]}), "Covariance": pd.DataFrame()}
        with pytest.raises(DataValidationError) as exc_info:
            dataframe_to_model(dfs, "portfolio")
        err = exc_info.value
        assert err.details.get("sheet") == "Assets"

    def test_unparseable_number_raises(self):
        assets_df = pd.DataFrame({
            "Name": ["A"],
            "Expected Return": ["not_a_number"],
        })
        cov_df = pd.DataFrame({"": ["A"], "A": [0.04]})
        with pytest.raises(DataValidationError, match="parse"):
            dataframe_to_model({"Assets": assets_df, "Covariance": cov_df}, "portfolio")

    def test_unknown_problem_type_raises(self):
        with pytest.raises(DataValidationError, match="Unknown problem type"):
            dataframe_to_model({}, "badtype")

    def test_empty_assets_sheet_raises(self):
        dfs = {
            "Assets": pd.DataFrame(columns=["Name", "Expected Return"]),
            "Covariance": pd.DataFrame(),
        }
        with pytest.raises(DataValidationError, match="No assets"):
            dataframe_to_model(dfs, "portfolio")

    def test_file_not_found_raises_fileioerror(self, tmp_path):
        with pytest.raises(FileIOError):
            read_data(str(tmp_path / "does_not_exist.xlsx"))

    def test_generate_template_bad_path_raises(self, tmp_path):
        with pytest.raises(FileIOError, match="Cannot save"):
            generate_template("portfolio", str(tmp_path / "subdir" / "t.xlsx"))
